#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Busca los 504 correos del Excel en Supabase por CEDULA (número de identificación),
no solo por correo. Así vemos si ya existen pero con correo/ciudad diferente.
"""
import io
import json
import os
import sys
import re
import urllib.error
import urllib.request
from openpyxl import load_workbook

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RUTA_ENV = os.path.join(PROYECTO_ROOT, ".env.local")
USER_AGENT = "panel-datos-etl/1.0"
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def cargar_env_local():
    if not os.path.isfile(RUTA_ENV):
        return
    with open(RUTA_ENV, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

class Supa:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def get_todo(self, ruta, page=1000):
        filas, offset = [], 0
        sep = "&" if "?" in ruta else "?"
        while True:
            headers = {
                "apikey": self.key,
                "Authorization": f"Bearer {self.key}",
                "User-Agent": USER_AGENT,
            }
            req = urllib.request.Request(
                f"{self.base}{ruta}{sep}limit={page}&offset={offset}",
                headers=headers,
            )
            try:
                with urllib.request.urlopen(req, timeout=120) as resp:
                    lote = json.loads(resp.read())
            except urllib.error.HTTPError as e:
                detalle = e.read().decode(errors="replace")[:500]
                raise RuntimeError(f"HTTP {e.code}: {detalle}") from None
            filas.extend(lote or [])
            if not lote or len(lote) < page:
                return filas
            offset += page

print("[1] Extrayendo cédulas del Excel...")
ruta_excel = r"C:\Users\EstudiantesJC\Downloads\Base Mr Bogotá.xlsx"
wb = load_workbook(ruta_excel, read_only=True, data_only=True)

cedulas_excel = {}  # cedula -> {nombre, correo, estado, pestaña}
for sheet_name in ["Nuevas 2026", "Antiguas"]:
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nombre = (row[2] or "").strip()
        correo = (row[3] or "").strip().lower()
        cedula_str = str(row[5] or "").strip()  # Columna F: Numero de identificacion
        estado = (row[0] or "").strip()

        if not cedula_str or cedula_str == "0":
            continue

        # Normalizar cédula (solo dígitos)
        cedula_norm = re.sub(r"\D", "", cedula_str)
        if not cedula_norm:
            continue

        if cedula_norm not in cedulas_excel:
            cedulas_excel[cedula_norm] = {
                "nombre": nombre,
                "correo": correo,
                "estado": estado,
                "pestaña": sheet_name,
            }

print(f"  ✓ Excel: {len(cedulas_excel)} cédulas únicas\n")

print("[2] Leyendo Supabase (TODOS los participants)...")
cargar_env_local()
url = os.environ.get("SUPABASE_URL", "")
key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not url or not key:
    print("ERROR: faltan credenciales Supabase")
    sys.exit(1)

supa = Supa(url, key)
todos_supa = supa.get_todo("/participants?select=id,nombre,email,cedula,ciudad,grupo_ciudad")

cedulas_supa = {}  # cedula -> {nombre, email, ciudad, id}
for p in todos_supa:
    cedula_str = str(p.get("cedula") or "").strip()
    cedula_norm = re.sub(r"\D", "", cedula_str)
    if not cedula_norm:
        continue

    email = (p.get("email") or "").strip().lower()
    nombre = (p.get("nombre") or "").strip()
    ciudad = (p.get("ciudad") or "").strip()

    if cedula_norm not in cedulas_supa:
        cedulas_supa[cedula_norm] = {
            "nombre": nombre,
            "email": email,
            "ciudad": ciudad,
            "id": p.get("id"),
        }

print(f"  ✓ Supabase: {len(cedulas_supa)} cédulas únicas\n")

# Cruce por cédula
print("[3] CRUZANDO POR CÉDULA:")
print("=" * 120)

encontradas = cedulas_excel.keys() & cedulas_supa.keys()
nuevas = cedulas_excel.keys() - cedulas_supa.keys()
solo_supa = cedulas_supa.keys() - cedulas_excel.keys()

print(f"\nYA EXISTEN EN SUPABASE (por cédula): {len(encontradas)}")
print("  Primeras 10:")
for cedula in sorted(encontradas)[:10]:
    excel_d = cedulas_excel[cedula]
    supa_d = cedulas_supa[cedula]
    match_correo = "✓" if excel_d["correo"] == supa_d["email"] else "✗"
    print(f"    {cedula:15} | Excel: {excel_d['correo']:40} | Supa: {supa_d['email']:40} | {match_correo}")
    if excel_d["correo"] != supa_d["email"]:
        print(f"                 ⚠️  Correos DIFERENTES: Excel {excel_d['nombre'][:30]} vs Supa {supa_d['nombre'][:30]}")

print(f"\nNUEVAS (solo en Excel, NO en Supabase): {len(nuevas)}")
print("  Primeras 10:")
for cedula in sorted(nuevas)[:10]:
    d = cedulas_excel[cedula]
    print(f"    {cedula:15} | {d['correo']:40} | {d['nombre'][:30]}")

print("\n" + "=" * 120)
print("ANÁLISIS DETALLADO:")
print("=" * 120)

# Analizar discrepancias
correos_iguales = 0
correos_diferentes = 0
ciudades_bogota_en_supa = 0

for cedula in encontradas:
    excel_d = cedulas_excel[cedula]
    supa_d = cedulas_supa[cedula]
    if excel_d["correo"] == supa_d["email"]:
        correos_iguales += 1
    else:
        correos_diferentes += 1
    if "BOGOTA" in supa_d["ciudad"].upper():
        ciudades_bogota_en_supa += 1

print(f"""
De las {len(encontradas)} personas que EXISTEN en Supabase por cédula:
  - {correos_iguales} tienen correo IDÉNTICO
  - {correos_diferentes} tienen correo DIFERENTE (cambiaron de correo)
  - {ciudades_bogota_en_supa} están registrados en Bogotá en Supabase

Interpretación:
  - Si 0 encontradas: Excel es completamente nuevo (no está en Supabase)
  - Si muchas encontradas pero correos diferentes: personas actualizaron correo
  - Si muchas encontradas: datos ya están en Supabase, solo necesita validación de correo actual

RESUMEN PARA CAMPAÑA:
  - Usar cédulas de {len(encontradas)} personas QUE YA EXISTEN en Supabase
  - Validar/actualizar sus correos actuales
  - Agregar {len(nuevas)} personas nuevas que NO estaban
  - Total para campaña: {len(cedulas_excel)} personas de Bogotá
""")

print(f"\nRESUMEN: total_excel={len(cedulas_excel)} encontradas_en_supa={len(encontradas)} nuevas_en_excel={len(nuevas)}")
