#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Compara: Supabase (215 Cundinamarca) vs Excel Base Mr Bogotá.xlsx (512 Bogotá)
Identifica: nuevos correos, omisiones, duplicados
"""
import io
import json
import os
import sys
import re
import urllib.error
import urllib.request
from openpyxl import load_workbook

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RUTA_ENV = os.path.join(PROYECTO_ROOT, ".env.local")
USER_AGENT = "panel-datos-etl/1.0"
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def cargar_env_local():
    if not os.path.isfile(RUTA_ENV):
        return
    with open(RUTA_ENV, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

class Supa:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def get_todo(self, ruta, page=1000):
        filas, offset = [], 0
        sep = "&" if "?" in ruta else "?"
        while True:
            headers = {
                "apikey": self.key,
                "Authorization": f"Bearer {self.key}",
                "User-Agent": USER_AGENT,
            }
            req = urllib.request.Request(
                f"{self.base}{ruta}{sep}limit={page}&offset={offset}",
                headers=headers,
            )
            try:
                with urllib.request.urlopen(req, timeout=120) as resp:
                    lote = json.loads(resp.read())
            except urllib.error.HTTPError as e:
                detalle = e.read().decode(errors="replace")[:500]
                raise RuntimeError(f"HTTP {e.code}: {detalle}") from None
            filas.extend(lote or [])
            if not lote or len(lote) < page:
                return filas
            offset += page

print("[1] Leyendo Supabase (Cundinamarca)...")
cargar_env_local()
url = os.environ.get("SUPABASE_URL", "")
key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not url or not key:
    print("ERROR: faltan credenciales Supabase")
    sys.exit(1)

supa = Supa(url, key)
todos_supa = supa.get_todo("/participants?select=id,nombre,email,ciudad")

correos_supa = set()
for p in todos_supa:
    ciudad = (p.get("ciudad") or "").strip().upper()
    if "BOGOTA" not in ciudad and "BOGOTA" not in (p.get("grupo_ciudad") or "").upper():
        continue
    correo = (p.get("email") or "").strip().lower()
    if correo and EMAIL_RE.match(correo):
        correos_supa.add(correo)

print(f"  ✓ Supabase Bogotá: {len(correos_supa)} correos")

print("\n[2] Leyendo Excel Base Mr Bogotá.xlsx...")
ruta_excel = r"C:\Users\EstudiantesJC\Downloads\Base Mr Bogotá.xlsx"
wb = load_workbook(ruta_excel, read_only=True, data_only=True)

correos_excel = {}  # correo -> {nombre, estado, fecha}
for sheet_name in ["Nuevas 2026", "Antiguas"]:
    ws = wb[sheet_name]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nombre = (row[2] or "").strip()  # Columna C: Nombre Completo
        correo = (row[3] or "").strip().lower()  # Columna D: Correo
        estado = (row[0] or "").strip()  # Columna A: Estado
        fecha = row[1]  # Columna B: Fecha Creación

        if not correo or not EMAIL_RE.match(correo):
            continue

        if correo not in correos_excel:
            correos_excel[correo] = {
                "nombre": nombre,
                "estado": estado,
                "fecha": fecha,
                "pestaña": sheet_name,
            }

print(f"  ✓ Excel: {len(correos_excel)} correos únicos")
print(f"    - Nuevas 2026: ~40 correos")
print(f"    - Antiguas: ~472 correos")

# Comparación
print("\n[3] COMPARACIÓN:")
print("=" * 100)

solo_excel = set(correos_excel.keys()) - correos_supa
solo_supa = correos_supa - set(correos_excel.keys())
ambos = set(correos_excel.keys()) & correos_supa

print(f"\nSOLO en Excel (NO en Supabase): {len(solo_excel)} correos")
print(f"  Ejemplos:")
for correo in sorted(solo_excel)[:10]:
    d = correos_excel[correo]
    print(f"    - {correo:40} | {d['nombre'][:30]:30} | {d['pestaña']:15}")
if len(solo_excel) > 10:
    print(f"    ... y {len(solo_excel) - 10} más")

print(f"\nSOLO en Supabase (NO en Excel): {len(solo_supa)} correos")
print(f"  (Probablemente de otros municipios que se filtraron mal, o datos anteriores a 2024)")

print(f"\nEN AMBOS (verificado): {len(ambos)} correos")

# Resumen
print("\n" + "=" * 100)
print("RESUMEN PARA PRÓXIMA CAMPAÑA:")
print("=" * 100)
print(f"""
Correos únicos por fuente:
  - Supabase (Bogotá solo):      {len(correos_supa)}
  - Excel (Base Mr Bogotá.xlsx): {len(correos_excel)}
  - Solapamiento (en ambas):     {len(ambos)}
  - Nuevos en Excel (no en Supa): {len(solo_excel)}

RECOMENDACIÓN:
Usar Excel como fuente de verdad para Bogotá 2026, ya que:
  1. Tiene datos más ricos (sociodemográficos, emprendimiento, ingresos)
  2. Tiene histórico desde 2024 (Supabase solo 2025/2026)
  3. Es la fuente original (formularios de MR)
  4. Cubre {len(correos_excel)} vs {len(correos_supa)} personas

ACCIÓN PENDIENTE:
  - Cargar este Excel a Supabase (tabla `postulantes_mr_historico` o similar)
  - Sincronizar regularmente (cada semana/mes)
  - Ver por qué no estaba en Supabase desde el inicio
""")

print(f"\nRESUMEN: supa={len(correos_supa)} excel={len(correos_excel)} nuevos={len(solo_excel)} solapamiento={len(ambos)}")
