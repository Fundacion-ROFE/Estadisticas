#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrae lista de Bogotá directamente del Excel (fuente de verdad).
504 personas, datos ricos, histórico 2024-2026.
Excluye rebotes si existen en Supabase.
"""
import io
import json
import os
import re
import sys
import urllib.error
import urllib.request
from openpyxl import load_workbook

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RUTA_ENV = os.path.join(PROYECTO_ROOT, ".env.local")
TOOLS_DATA = os.path.join(PROYECTO_ROOT, "tools", "mujeres-rofe-correos", "data")
RUTA_SALIDA = os.path.join(TOOLS_DATA, "lista_bogota_excel_2026_jul24.csv")

USER_AGENT = "panel-datos-etl/1.0"
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def log(msg):
    print(f"[bogota-excel] {msg}", flush=True)

def cargar_env_local():
    if not os.path.isfile(RUTA_ENV):
        return
    with open(RUTA_ENV, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

class Supa:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def get_todo(self, ruta, page=1000):
        filas, offset = [], 0
        sep = "&" if "?" in ruta else "?"
        while True:
            headers = {
                "apikey": self.key,
                "Authorization": f"Bearer {self.key}",
                "User-Agent": USER_AGENT,
            }
            req = urllib.request.Request(
                f"{self.base}{ruta}{sep}limit={page}&offset={offset}",
                headers=headers,
            )
            try:
                with urllib.request.urlopen(req, timeout=120) as resp:
                    lote = json.loads(resp.read())
            except urllib.error.HTTPError as e:
                detalle = e.read().decode(errors="replace")[:500]
                raise RuntimeError(f"HTTP {e.code}: {detalle}") from None
            filas.extend(lote or [])
            if not lote or len(lote) < page:
                return filas
            offset += page

def extraer_supresiones():
    """Correos en email_bounces (hard) + email_optout."""
    cargar_env_local()
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        return set(), 0, 0
    supa = Supa(url, key)
    optout, hard = set(), set()
    try:
        optout = {(f.get("email") or "").strip().lower()
                  for f in supa.get_todo("/email_optout?select=email") if f.get("email")}
    except RuntimeError as e:
        log(f"AVISO: no se pudo leer email_optout ({e})")
    try:
        hard = {(f.get("email") or "").strip().lower()
                for f in supa.get_todo("/email_bounces?select=email&tipo=eq.hard") if f.get("email")}
    except RuntimeError as e:
        log(f"AVISO: no se pudo leer email_bounces ({e})")
    return optout | hard, len(optout), len(hard)

def main():
    log("Leyendo Excel 'Base Mr Bogotá.xlsx'...")
    ruta_excel = r"C:\Users\EstudiantesJC\Downloads\Base Mr Bogotá.xlsx"
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)

    mapa = {}
    for sheet_name in ["Nuevas 2026", "Antiguas"]:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            nombre = (row[2] or "").strip()  # Nombre Completo
            correo = (row[3] or "").strip().lower()  # Correo
            estado = (row[0] or "").strip()  # Estado

            if not correo or not EMAIL_RE.match(correo):
                continue

            mapa[correo] = {
                "nombre": nombre,
                "estado": estado,
                "pestaña": sheet_name,
            }

    log(f"  Encontrados: {len(mapa)} correos únicos")

    # Excluir supresiones
    supresiones, n_optout, n_hard = extraer_supresiones()
    excluidos = 0
    if supresiones:
        antes = len(mapa)
        mapa = {c: d for c, d in mapa.items() if c not in supresiones}
        excluidos = antes - len(mapa)

    log(f"Supresiones: {n_optout} opt-out, {n_hard} hard bounces → {excluidos} excluidos")
    log(f"Lista final: {len(mapa)} correos")

    # Generar CSV
    os.makedirs(TOOLS_DATA, exist_ok=True)
    with open(RUTA_SALIDA, "w", encoding="utf-8", newline="") as f:
        f.write("nombre,correo,cohorte\n")
        for correo in sorted(mapa.keys()):
            d = mapa[correo]
            nombre = (d.get("nombre") or "").replace('"', '""')
            # cohorte = pestaña (Nuevas 2026 o Antiguas)
            cohorte = d.get("pestaña", "")
            f.write(f'"{nombre}",{correo},{cohorte}\n')

    log(f"✓ Salida: {RUTA_SALIDA}")
    print(f"RESUMEN: bogota_excel={len(mapa)} excluidos={excluidos} estado=exito")

if __name__ == "__main__":
    main()
