#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDITORÍA AUTOMÁTICA: ¿Por qué 504 personas de Bogotá no están en Supabase?

Ejecutar: python auditar_datos_faltantes.py
"""
import io
import json
import os
import sys
import re
import urllib.error
import urllib.request
from openpyxl import load_workbook
from collections import defaultdict

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RUTA_ENV = os.path.join(PROYECTO_ROOT, ".env.local")
USER_AGENT = "panel-datos-etl/1.0"
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def cargar_env_local():
    if not os.path.isfile(RUTA_ENV):
        return
    with open(RUTA_ENV, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())

class Supa:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def get_todo(self, ruta, page=1000):
        filas, offset = [], 0
        sep = "&" if "?" in ruta else "?"
        while True:
            headers = {
                "apikey": self.key,
                "Authorization": f"Bearer {self.key}",
                "User-Agent": USER_AGENT,
            }
            req = urllib.request.Request(
                f"{self.base}{ruta}{sep}limit={page}&offset={offset}",
                headers=headers,
            )
            try:
                with urllib.request.urlopen(req, timeout=120) as resp:
                    lote = json.loads(resp.read())
            except urllib.error.HTTPError as e:
                detalle = e.read().decode(errors="replace")[:500]
                raise RuntimeError(f"HTTP {e.code}: {detalle}") from None
            filas.extend(lote or [])
            if not lote or len(lote) < page:
                return filas
            offset += page

print("╔" + "═" * 98 + "╗")
print("║ AUDITORÍA: ¿POR QUÉ FALTAN 504 PERSONAS DE BOGOTÁ EN SUPABASE?".ljust(99) + "║")
print("╚" + "═" * 98 + "╝\n")

# PASO 1: Extraer correos del Excel
print("[PASO 1] Extrayendo correos del Excel...")
ruta_excel = r"C:\Users\EstudiantesJC\Downloads\Base Mr Bogotá.xlsx"
correos_excel = {}
try:
    wb = load_workbook(ruta_excel, read_only=True, data_only=True)
    for sheet_name in ["Nuevas 2026", "Antiguas"]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre = (row[2] or "").strip()
            correo = (row[3] or "").strip().lower()
            if correo and EMAIL_RE.match(correo):
                correos_excel[correo] = {"nombre": nombre, "pestaña": sheet_name}
except Exception as e:
    print(f"❌ ERROR: {e}")
    sys.exit(1)

print(f"  ✓ {len(correos_excel)} correos únicos en Excel\n")

# PASO 2: Conectar a Supabase y revisar estructura
print("[PASO 2] Conectando a Supabase...")
cargar_env_local()
url = os.environ.get("SUPABASE_URL", "")
key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not url or not key:
    print("❌ ERROR: faltan credenciales Supabase en .env.local")
    sys.exit(1)

supa = Supa(url, key)

# Paso 2a: Revisar si existen tablas alternativas
print("  Consultando tablas disponibles...")
try:
    info_schema = supa.get_todo("/information_schema.tables?table_schema=eq.public&select=table_name")
    tablas = [t.get("table_name") for t in info_schema]
    tablas_mr = [t for t in tablas if "postulante" in t.lower() or "mr" in t.lower()]
    print(f"  ✓ Tablas encontradas: {len(tablas)}")
    if tablas_mr:
        print(f"    → Tablas MR/postulantes: {tablas_mr}")
except Exception as e:
    print(f"  ⚠ No se pudieron listar tablas: {e}")
    tablas_mr = []

# PASO 3: Buscar correos en participants (SIN filtro ciudad)
print("\n[PASO 3] Buscando correos en participants (TODOS, sin filtro)...")
todos_participants = supa.get_todo("/participants?select=id,email,nombre,ciudad,grupo_ciudad,source_system")
correos_supa = {}
ciudades_encontradas = defaultdict(int)

for p in todos_participants:
    correo = (p.get("email") or "").strip().lower()
    if not correo:
        continue
    correos_supa[correo] = {
        "nombre": p.get("nombre"),
        "ciudad": p.get("ciudad"),
        "grupo_ciudad": p.get("grupo_ciudad"),
        "source_system": p.get("source_system"),
        "id": p.get("id"),
    }
    ciudad = (p.get("ciudad") or "").strip()
    if ciudad:
        ciudades_encontradas[ciudad] += 1

print(f"  ✓ {len(correos_supa)} correos en participants")

# PASO 4: Cruzar correos Excel vs Supabase
encontrados = set(correos_excel.keys()) & set(correos_supa.keys())
no_encontrados = set(correos_excel.keys()) - set(correos_supa.keys())
solo_supa = set(correos_supa.keys()) - set(correos_excel.keys())

print("\n[PASO 4] CRUZANDO POR CORREO:")
print(f"  ✓ ENCONTRADOS en Supabase: {len(encontrados)} / {len(correos_excel)}")
print(f"  ✗ NO encontrados: {len(no_encontrados)} / {len(correos_excel)}")
print(f"  → Cobertura: {100*len(encontrados)//len(correos_excel)}%\n")

# Mostrar ejemplos de encontrados
if encontrados:
    print("  Ejemplos ENCONTRADOS en Supabase:")
    for correo in sorted(encontrados)[:5]:
        d = correos_supa[correo]
        print(f"    ✓ {correo:40} | Ciudad: {(d['ciudad'] or 'NULL'):20} | Source: {d['source_system']}")
    if len(encontrados) > 5:
        print(f"    ... y {len(encontrados)-5} más")

# Mostrar ejemplos de no encontrados
if no_encontrados:
    print(f"\n  Ejemplos NO encontrados en Supabase:")
    for correo in sorted(no_encontrados)[:5]:
        d = correos_excel[correo]
        print(f"    ✗ {correo:40} | {d['nombre'][:30]:30} | Pestaña: {d['pestaña']}")
    if len(no_encontrados) > 5:
        print(f"    ... y {len(no_encontrados)-5} más")

# PASO 5: Análisis de variantes de ciudad
print("\n[PASO 5] VARIANTES DE CIUDAD EN SUPABASE:")
bogota_variantes = {k: v for k, v in ciudades_encontradas.items() if "BOGOTA" in k.upper()}
print(f"  Variantes detectadas:")
for ciudad in sorted(bogota_variantes.keys(), key=lambda c: -bogota_variantes[c]):
    print(f"    → {ciudad:30} : {bogota_variantes[ciudad]:4} personas")

# PASO 6: Revisar source_system
print("\n[PASO 6] SOURCE_SYSTEM en Supabase:")
source_systems = defaultdict(int)
for p in todos_participants:
    ss = p.get("source_system") or "(none)"
    source_systems[ss] += 1
for ss in sorted(source_systems.keys()):
    print(f"  {ss:20} : {source_systems[ss]:6} participants")

# PASO 7: Revisar postulantes_mr si existe
postulantes_mr_count = 0
if "postulantes_mr" in tablas_mr or "postulantes_mr" in tablas:
    print("\n[PASO 7] REVISANDO TABLA postulantes_mr:")
    try:
        postulantes_mr = supa.get_todo("/postulantes_mr?select=email,ciudad&limit=10")
        postulantes_mr_count = len(postulantes_mr)

        # Buscar correos Excel en postulantes_mr (muestra)
        correos_pm = {(p.get("email") or "").lower() for p in postulantes_mr}
        en_pm = set(correos_excel.keys()) & correos_pm
        print(f"  Registros en postulantes_mr: {postulantes_mr_count} (muestra de 10)")
        print(f"  Correos Excel encontrados en postulantes_mr: {len(en_pm)}")
        if en_pm:
            print(f"    Ejemplos: {list(en_pm)[:3]}")
    except Exception as e:
        print(f"  ⚠ Error consultando postulantes_mr: {e}")

# PASO 8: Reporte final
print("\n" + "=" * 100)
print("REPORTE DE AUDITORÍA")
print("=" * 100)

reporte = f"""
┌─ DATOS FALTANTES ────────────────────────────────────────────────────┐
│ Excel (Base Mr Bogotá.xlsx):  {len(correos_excel):4} personas                       │
│ Supabase (participants):      {len(correos_supa):4} personas                       │
│ Encontrados por correo:       {len(encontrados):4} personas ({100*len(encontrados)//len(correos_excel):3}%)                    │
│ NO encontrados:               {len(no_encontrados):4} personas ({100*len(no_encontrados)//len(correos_excel):3}%) ← PROBLEMA │
└──────────────────────────────────────────────────────────────────────┘

┌─ DÓNDE PODRÍAN ESTAR ────────────────────────────────────────────────┐
│ ✓ En participants (encontrados):          {len(encontrados):4}          │
│ ? En postulantes_mr (si existe):          ~???           │
│ ? En enrollments MR (como matrículas):    ~???           │
│ ✗ NO están en Supabase (confirmado):      {len(no_encontrados):4}          │
└──────────────────────────────────────────────────────────────────────┘

┌─ POSIBLES CAUSAS ────────────────────────────────────────────────────┐
│ 1. Normalización de ciudad fallida (Bogotá D.C. vs Bogota vs BOGOTA) │
│ 2. Los datos están en participants pero source_system ≠ 'mr'         │
│ 3. Los datos nunca se cargaron de MongoDB (falta de sincronización)  │
│ 4. Los datos están en otra tabla (postulantes_mr no revisada)        │
│ 5. Los datos están como enrollments, no como base participants       │
└──────────────────────────────────────────────────────────────────────┘

┌─ RECOMENDACIÓN ──────────────────────────────────────────────────────┐
│ PRÓXIMO PASO: Ejecutar consultas SQL en Supabase (ver abajo)        │
│              para determinar dónde están realmente los datos         │
└──────────────────────────────────────────────────────────────────────┘
"""
print(reporte)

# PASO 9: Queries SQL sugeridas
print("\n" + "=" * 100)
print("CONSULTAS SQL PARA INVESTIGAR (ejecutar en Supabase SQL Editor):")
print("=" * 100)
print("""
-- [Q1] ¿Cuántos registros en postulantes_mr?
SELECT COUNT(*) as total,
       COUNT(DISTINCT UPPER(ciudad)) as ciudades_unicas
FROM postulantes_mr;

-- [Q2] Variantes de ciudad en postulantes_mr (Bogotá)
SELECT DISTINCT ciudad, COUNT(*) as qty
FROM postulantes_mr
WHERE ciudad ILIKE '%bogot%'
GROUP BY ciudad
ORDER BY qty DESC;

-- [Q3] ¿Están estos 5 correos en postulantes_mr?
SELECT email, ciudad FROM postulantes_mr
WHERE email IN (
  'damilsav72@gmail.com',
  'angelacepedaadm@gmail.com',
  'peralta.lit2107@gmail.com',
  'sipitoca@hotmail.com',
  'ymartinez.cedi@gmail.com'
);

-- [Q4] Participants con source_system='mr' y Bogotá
SELECT COUNT(*), ciudad FROM participants
WHERE source_system = 'mr' AND ciudad ILIKE '%bogot%'
GROUP BY ciudad;

-- [Q5] ¿Hay otros source_system para MR?
SELECT DISTINCT source_system, COUNT(*) as qty
FROM participants p
JOIN enrollments e ON p.id = e.participant_id
JOIN courses c ON e.course_id = c.id
WHERE c.programa = 'mr'
GROUP BY source_system;

-- [Q6] Participants MR de Bogotá (sin duplicar por filtro)
SELECT COUNT(DISTINCT p.id) as participants_mr_bogota
FROM participants p
JOIN enrollments e ON p.id = e.participant_id
JOIN courses c ON e.course_id = c.id
WHERE c.programa = 'mr'
  AND (p.ciudad ILIKE '%bogot%' OR p.grupo_ciudad ILIKE '%bogot%');
""")

print("\n" + "=" * 100)
print("✋ SUSPENDIDO AQUÍ — Revisa las consultas SQL arriba")
print("=" * 100)
print("""
1. Copia las queries de arriba
2. Ve a Supabase SQL Editor: https://supabase.com/projects/...
3. Ejecuta Q1, Q2, Q3, Q4, Q5, Q6
4. Reporta resultados aquí
5. Basándote en eso, sabremos dónde están los datos
""")
