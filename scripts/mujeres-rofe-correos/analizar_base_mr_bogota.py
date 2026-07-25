#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Análisis rápido: "Base Mr Bogotá.xlsx" — qué datos hay y qué no está en Supabase.
"""
import sys
from openpyxl import load_workbook
from collections import Counter

ruta = r"C:\Users\EstudiantesJC\Downloads\Base Mr Bogotá.xlsx"

try:
    wb = load_workbook(ruta, read_only=True, data_only=True)
except Exception as e:
    print(f"ERROR al abrir {ruta}: {e}")
    sys.exit(1)

print(f"[*] Leyendo: {ruta}\n")
print(f"[*] Pestañas disponibles: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"{'='*100}")
    print(f"PESTAÑA: {sheet_name}")
    print(f"{'='*100}")

    # Leer encabezados (primera fila)
    header = []
    for cell in ws[1]:
        header.append(cell.value or "")
    print(f"Columnas ({len(header)}): {header}\n")

    # Contar filas
    row_count = 0
    correos = []
    ciudades = []
    nombres = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_count += 1

        # Busca correo (común en varias columnas)
        for col_idx, val in enumerate(row):
            if val and "@" in str(val):
                correos.append(str(val).strip().lower())

        # Busca ciudad (probablemente hay una columna)
        # Por ahora, mostrar primeras filas
        if row_idx <= 5:
            print(f"  Fila {row_idx}: {row}")

    print(f"\nTotal filas (datos): {row_count}")
    print(f"Correos encontrados: {len(set(correos))}")
    if correos:
        print(f"  Primeros 5: {list(set(correos))[:5]}")

    print()

print(f"\n[*] Análisis COMPLETADO")
