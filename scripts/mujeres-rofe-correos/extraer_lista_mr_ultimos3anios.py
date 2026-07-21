#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extraer_lista_mr_ultimos3anios.py — Lista de envío Mujeres ROFÉ (últimos 3 años).

Combina dos fuentes porque Supabase (panel-datos-rofe) AÚN NO tiene el histórico
completo del programa MR: `courses.programa='mr'` solo existe para cohortes 2025 y
2026 (no hay matrículas MR cargadas para 2023/2024 — ver [[project-panel-datos-supabase]]).

  1. Supabase: participants↔enrollments↔courses WHERE programa='mr' (cohortes que
     existan realmente en la base — hoy 2025/2026).
  2. Excel BD-Mujeres ROFÉ (pestaña General): columna "Fecha de Creación" (c3),
     filtrando año en los últimos 3 años calendario. Cubre 2024, que Supabase no tiene.

Unión por correo (lowercase), Supabase gana si el correo aparece en ambas fuentes
(tiene más contexto: cohorte real de curso). Se marca la fuente de cada fila.

⚠ PRIVACIDAD: la salida (nombre + correo real) se escribe SOLO en tools/ (gitignoreado).
Este script no debe escribir CSVs con PII bajo scripts/ (ver CLAUDE.md).

⚠ PENDIENTE (anotado, no resuelto aquí): para que Supabase "tenga noción de todo"
como pide el equipo, falta un import histórico MR 2023-2024 análogo a
importar_historico_q10.py (que sí se hizo para JC). Este script es un parche de
lectura, no reemplaza esa carga.

Uso:
    python extraer_lista_mr_ultimos3anios.py [--ruta-bd "C:\\...\\BD-Mujeres ROFÉ ....xlsx"]
Al final excluye supresiones de Supabase: `email_optout` (baja voluntaria) + `email_bounces`
con tipo=hard (rebotes permanentes). Los soft bounces (temporales) NO se excluyen.

Consola (parseable):
    RESUMEN: supabase=N excel=M union=T solapados=O suprimidos=X estado=exito
"""
import argparse
import io
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RUTA_ENV = os.path.join(PROYECTO_ROOT, ".env.local")
RUTA_BD_DEFAULT = r"C:\Users\EstudiantesJC\Downloads\BD-Mujeres ROFÉ 2026 (2).xlsx"
TOOLS_DATA = os.path.join(PROYECTO_ROOT, "tools", "mujeres-rofe-correos", "data")
RUTA_SALIDA = os.path.join(TOOLS_DATA, "lista_mr_ultimos_3_anios.csv")

USER_AGENT = "panel-datos-etl/1.0"
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def log(msg):
    print(f"[extraer-mr] {msg}", flush=True)


def cargar_env_local():
    if not os.path.isfile(RUTA_ENV):
        return
    with open(RUTA_ENV, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())


class Supa:
    def __init__(self, url, key):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def get_todo(self, ruta, page=1000):
        filas, offset = [], 0
        sep = "&" if "?" in ruta else "?"
        while True:
            headers = {
                "apikey": self.key,
                "Authorization": f"Bearer {self.key}",
                "User-Agent": USER_AGENT,
            }
            req = urllib.request.Request(
                f"{self.base}{ruta}{sep}limit={page}&offset={offset}",
                headers=headers,
            )
            try:
                with urllib.request.urlopen(req, timeout=120) as resp:
                    lote = json.loads(resp.read())
            except urllib.error.HTTPError as e:
                detalle = e.read().decode(errors="replace")[:500]
                raise RuntimeError(f"HTTP {e.code} en GET {ruta}: {detalle}") from None
            filas.extend(lote or [])
            if not lote or len(lote) < page:
                return filas
            offset += page


def extraer_supabase():
    """Devuelve dict correo_lower -> {nombre, cohorte, fuente='supabase'}."""
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        log("AVISO: faltan credenciales Supabase (.env.local) — se omite esta fuente")
        return {}

    supa = Supa(url, key)
    cursos = supa.get_todo("/courses?programa=eq.mr&select=id,cohorte")
    if not cursos:
        log("AVISO: no hay courses con programa=mr en Supabase")
        return {}
    cohorte_por_curso = {c["id"]: c["cohorte"] for c in cursos}
    ids_cursos = list(cohorte_por_curso.keys())

    matriculas = []
    LOTE_IN = 200
    for i in range(0, len(ids_cursos), LOTE_IN):
        grupo = ids_cursos[i:i + LOTE_IN]
        filtro = ",".join(grupo)
        matriculas.extend(supa.get_todo(
            f"/enrollments?course_id=in.({filtro})&select=participant_id,course_id"))

    cohorte_por_participante = {}
    for m in matriculas:
        cohorte = cohorte_por_curso.get(m["course_id"])
        actual = cohorte_por_participante.get(m["participant_id"])
        if actual is None or (cohorte and cohorte > actual):
            cohorte_por_participante[m["participant_id"]] = cohorte

    ids_participantes = list(cohorte_por_participante.keys())
    resultado = {}
    for i in range(0, len(ids_participantes), LOTE_IN):
        grupo = ids_participantes[i:i + LOTE_IN]
        filtro = ",".join(grupo)
        participantes = supa.get_todo(
            f"/participants?id=in.({filtro})&select=id,nombre,email")
        for p in participantes:
            correo = (p.get("email") or "").strip().lower()
            if not correo or not EMAIL_RE.match(correo):
                continue
            resultado[correo] = {
                "nombre": p.get("nombre") or "",
                "cohorte": cohorte_por_participante.get(p["id"], ""),
                "fuente": "supabase",
            }
    return resultado


def extraer_supresiones():
    """Set de correos (lowercase) a excluir: opt-out (baja voluntaria) + hard bounces
    (rebotes permanentes de email_bounces, tipo=hard). Los soft bounces NO se excluyen
    (son temporales). Vacío si no hay credenciales o las tablas no responden."""
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        return set(), 0, 0
    supa = Supa(url, key)
    optout, hard = set(), set()
    try:
        optout = {(f.get("email") or "").strip().lower()
                  for f in supa.get_todo("/email_optout?select=email") if f.get("email")}
    except RuntimeError as e:
        log(f"AVISO: no se pudo leer email_optout ({e}) — no se excluye opt-out")
    try:
        hard = {(f.get("email") or "").strip().lower()
                for f in supa.get_todo("/email_bounces?select=email&tipo=eq.hard") if f.get("email")}
    except RuntimeError as e:
        log(f"AVISO: no se pudo leer email_bounces ({e}) — no se excluyen rebotes")
    return optout | hard, len(optout), len(hard)


def year_of(fecha):
    if fecha is None:
        return None
    if hasattr(fecha, "year"):
        return fecha.year
    s = str(fecha).strip()
    return int(s[:4]) if s[:4].isdigit() else None


def extraer_excel(ruta_bd, anios_objetivo):
    """Devuelve dict correo_lower -> {nombre, cohorte, fuente='excel'}."""
    if not os.path.isfile(ruta_bd):
        log(f"AVISO: no se encontró el Excel BD-Mujeres ROFÉ en {ruta_bd} — se omite esta fuente")
        return {}

    wb = load_workbook(ruta_bd, read_only=True, data_only=True)
    ws = wb["General"]
    resultado = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha = row[2]
        anio = year_of(fecha)
        if anio not in anios_objetivo:
            continue
        nombre = (row[3] or "").strip() if row[3] else ""
        correo = (row[4] or row[7] or "").strip().lower()
        if not correo or not EMAIL_RE.match(correo):
            continue
        # Si hay duplicado dentro del Excel, se queda con la fecha más reciente.
        existente = resultado.get(correo)
        if existente is None or str(anio) >= existente["cohorte"]:
            resultado[correo] = {"nombre": nombre, "cohorte": str(anio), "fuente": "excel"}
    return resultado


def main():
    ap = argparse.ArgumentParser(description="Extrae lista de envío MR (últimos 3 años)")
    ap.add_argument("--ruta-bd", default=RUTA_BD_DEFAULT, help="Ruta al Excel BD-Mujeres ROFÉ")
    args = ap.parse_args()

    cargar_env_local()

    anio_actual = datetime.now().year
    anios_objetivo = {anio_actual, anio_actual - 1, anio_actual - 2}
    log(f"Últimos 3 años calendario: {sorted(anios_objetivo)}")

    log("Consultando Supabase (programa=mr)...")
    mapa_supabase = extraer_supabase()
    log(f"  Supabase: {len(mapa_supabase)} correos únicos "
        f"(cohortes disponibles: {sorted(set(v['cohorte'] for v in mapa_supabase.values()))})")

    log(f"Leyendo Excel BD-Mujeres ROFÉ ({args.ruta_bd})...")
    mapa_excel = extraer_excel(args.ruta_bd, anios_objetivo)
    log(f"  Excel: {len(mapa_excel)} correos únicos en {sorted(anios_objetivo)}")

    solapados = set(mapa_supabase) & set(mapa_excel)
    union = dict(mapa_excel)
    union.update(mapa_supabase)  # Supabase gana en solapamiento (más contexto)

    # Excluir supresiones: opt-out (baja voluntaria) + hard bounces (rebotes permanentes).
    supresiones, n_optout, n_hard = extraer_supresiones()
    excluidos = 0
    if supresiones:
        antes = len(union)
        union = {c: d for c, d in union.items() if c not in supresiones}
        excluidos = antes - len(union)
        log(f"Supresiones: {n_optout} opt-out + {n_hard} hard bounces = "
            f"{len(supresiones)} únicos → {excluidos} excluidos de la unión")

    os.makedirs(TOOLS_DATA, exist_ok=True)
    with open(RUTA_SALIDA, "w", encoding="utf-8-sig", newline="") as f:
        f.write("nombre,correo,cohorte,fuente\n")
        for correo, datos in sorted(union.items()):
            nombre = datos["nombre"].replace('"', "'")
            f.write(f'"{nombre}",{correo},{datos["cohorte"]},{datos["fuente"]}\n')

    log(f"Escrito: {RUTA_SALIDA}")
    print(f"RESUMEN: supabase={len(mapa_supabase)} excel={len(mapa_excel)} "
          f"union={len(union)} solapados={len(solapados)} suprimidos={excluidos} "
          f"estado=exito")


if __name__ == "__main__":
    main()
