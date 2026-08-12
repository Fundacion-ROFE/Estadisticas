# -*- coding: utf-8 -*-
"""
catalogar_fuentes_historicas.py — Reduce un árbol de Excel/CSV gigantes a .md navegables.

Motivación: la etapa final de alimentación de la DB requiere leer decenas de bases
históricas (JC 2019-2025 + Mujeres ROFÉ) en .xlsx/.csv de hasta ~27 MB. Cargarlas en
un LLM es inviable; este script las recorre en streaming (openpyxl read_only / csv
streaming — memoria plana aunque el archivo pese) y emite por cada archivo un .md con:
  - hojas, dimensiones (filas x columnas)
  - por columna: no-nulos (n y %), valores únicos, dtype inferido, muestras
  - primeras filas de muestra
Además un INDICE.md global y un catalogo.json legible por máquina (para el análisis
de brechas y para configurar los subagentes de extracción).

PII: la salida contiene datos personales (nombres, correos, cédulas) → SIEMPRE va a
`tools/` (gitignoreado). NUNCA a docs/ ni al repo público.

Uso:
    python catalogar_fuentes_historicas.py \
        --root "C:/Users/EstudiantesJC/Downloads/COMPLETE-ORDEN-INFORMATION" \
        --salida "../../tools/catalogo_complete_orden"
    # opcional: --muestras 6  --max-cols 120
"""

import argparse
import csv
import json
import os
import re
import sys
from collections import Counter
from datetime import date, datetime

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = __import__("io").TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

EXT_EXCEL = {".xlsx", ".xlsm", ".xls"}
EXT_CSV = {".csv"}
MUESTRA_MAX_LEN = 60          # trunca celdas largas en las muestras
FILAS_ESCANEO_HEADER = 15     # filas iniciales para detectar la fila de encabezado


def log(msg):
    print(f"[catalogo] {msg}", flush=True)


def _norm(v):
    """Normaliza una celda a str corto para muestra/heurística."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    s = str(v).strip()
    return s


def _dtype_de(valores):
    """Infiere un dtype grosero a partir de una lista de valores no vacíos (muestra)."""
    if not valores:
        return "vacio"
    n_int = n_float = n_fecha = n_email = 0
    for v in valores:
        s = str(v).strip()
        if isinstance(v, (datetime, date)):
            n_fecha += 1
            continue
        if re.fullmatch(r"-?\d+", s):
            n_int += 1
        elif re.fullmatch(r"-?\d+[.,]\d+", s):
            n_float += 1
        if re.search(r"[^@\s]+@[^@\s]+\.[^@\s]+", s):
            n_email += 1
    total = len(valores)
    if n_email / total > 0.5:
        return "email"
    if n_fecha / total > 0.5:
        return "fecha"
    if (n_int + n_float) / total > 0.8:
        return "numerico" if n_float else "entero"
    return "texto"


def _detectar_header(filas_muestra):
    """Devuelve el índice (0-based) de la fila que mejor parece encabezado:
    la de mayor cantidad de celdas de texto corto y no vacías dentro del escaneo."""
    mejor_i, mejor_score = 0, -1
    for i, fila in enumerate(filas_muestra):
        celdas = [_norm(c) for c in fila]
        no_vacias = [c for c in celdas if c]
        cortas = [c for c in no_vacias if len(c) <= MUESTRA_MAX_LEN and not re.fullmatch(r"-?\d+([.,]\d+)?", c)]
        score = len(cortas) + 0.3 * len(no_vacias)
        if score > mejor_score:
            mejor_i, mejor_score = i, score
    return mejor_i


def _cols_unicas(nombres):
    """Desambigua encabezados repetidos/vacíos: col, col_2, col_3, (vacio_N)."""
    vistos = Counter()
    salida = []
    for idx, n in enumerate(nombres):
        base = n if n else f"(col_{idx+1}_sin_nombre)"
        vistos[base] += 1
        salida.append(base if vistos[base] == 1 else f"{base}__{vistos[base]}")
    return salida


def procesar_hoja(nombre_hoja, iter_filas, muestras, max_cols):
    """Recorre una hoja fila-a-fila (streaming). Devuelve dict con estructura + stats."""
    buffer_ini = []
    for fila in iter_filas:
        buffer_ini.append(fila)
        if len(buffer_ini) >= FILAS_ESCANEO_HEADER:
            break

    if not buffer_ini:
        return {"nombre": nombre_hoja, "filas_datos": 0, "columnas": [], "muestras": [], "vacia": True}

    h_idx = _detectar_header(buffer_ini)
    header = [_norm(c) for c in buffer_ini[h_idx]]
    header = _cols_unicas(header)
    if max_cols and len(header) > max_cols:
        header = header[:max_cols]
    ncols = len(header)

    no_nulos = [0] * ncols
    unicos = [set() for _ in range(ncols)]
    muestra_vals = [[] for _ in range(ncols)]
    filas_muestra = []
    n_datos = 0

    # Las filas posteriores al header dentro del buffer inicial son datos ya leídos
    resto_buffer = buffer_ini[h_idx + 1:]

    def _consumir(fila):
        nonlocal n_datos
        celdas = list(fila)[:ncols]
        if all(_norm(c) == "" for c in celdas):
            return  # fila totalmente vacía (grilla de Google Forms) → no cuenta como dato
        n_datos += 1
        for j in range(ncols):
            v = celdas[j] if j < len(celdas) else None
            s = _norm(v)
            if s != "":
                no_nulos[j] += 1
                if len(unicos[j]) < 5000:
                    unicos[j].add(s[:MUESTRA_MAX_LEN])
                if len(muestra_vals[j]) < 8:
                    muestra_vals[j].append(v)
        if len(filas_muestra) < muestras:
            filas_muestra.append([_norm(c)[:MUESTRA_MAX_LEN] for c in celdas])

    for fila in resto_buffer:
        _consumir(fila)
    for fila in iter_filas:
        _consumir(fila)

    columnas = []
    for j, nombre in enumerate(header):
        pct = round(100.0 * no_nulos[j] / n_datos, 1) if n_datos else 0.0
        columnas.append({
            "nombre": nombre,
            "no_nulos": no_nulos[j],
            "pct_lleno": pct,
            "n_unicos": len(unicos[j]),
            "dtype": _dtype_de([v for v in muestra_vals[j]]),
            "muestras": [_norm(v)[:MUESTRA_MAX_LEN] for v in muestra_vals[j][:5]],
        })

    return {
        "nombre": nombre_hoja,
        "filas_datos": n_datos,
        "n_columnas": ncols,
        "columnas": columnas,
        "muestras": filas_muestra,
        "vacia": n_datos == 0,
    }


def procesar_excel(path, muestras, max_cols):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    hojas = []
    for ws in wb.worksheets:
        try:
            it = ws.iter_rows(values_only=True)
            hojas.append(procesar_hoja(ws.title, it, muestras, max_cols))
        except Exception as e:  # una hoja corrupta no debe tumbar el archivo
            hojas.append({"nombre": ws.title, "error": str(e)})
    wb.close()
    return hojas


def procesar_csv(path, muestras, max_cols):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as fh:
                muestra = fh.read(4096)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
                    sep = dialect.delimiter
                except Exception:
                    sep = ","
                reader = csv.reader(fh, delimiter=sep)
                hoja = procesar_hoja("(csv)", reader, muestras, max_cols)
                hoja["separador"] = sep
                hoja["encoding"] = enc
                return [hoja]
        except UnicodeDecodeError:
            continue
    return [{"nombre": "(csv)", "error": "no se pudo decodificar"}]


def md_de_archivo(rel, meta):
    L = [f"# {rel}", ""]
    L.append(f"- **Tamaño:** {meta['tamano_mb']} MB · **Tipo:** {meta['ext']} · **Hojas:** {len(meta['hojas'])}")
    L.append("")
    for h in meta["hojas"]:
        if h.get("error"):
            L.append(f"## Hoja `{h['nombre']}` — ⚠️ ERROR: {h['error']}"); L.append(""); continue
        if h.get("vacia"):
            L.append(f"## Hoja `{h['nombre']}` — (vacía)"); L.append(""); continue
        extra = ""
        if h.get("separador"):
            extra = f" · sep=`{h['separador']}` · enc=`{h['encoding']}`"
        L.append(f"## Hoja `{h['nombre']}` — {h['filas_datos']} filas × {h['n_columnas']} columnas{extra}")
        L.append("")
        L.append("| # | Columna | Lleno | Únicos | Tipo | Muestras |")
        L.append("|---|---|---|---|---|---|")
        for i, c in enumerate(h["columnas"], 1):
            muestras = " · ".join(m for m in c["muestras"] if m)[:120].replace("|", "\\|")
            nombre = c["nombre"].replace("|", "\\|")
            L.append(f"| {i} | {nombre} | {c['no_nulos']} ({c['pct_lleno']}%) | {c['n_unicos']} | {c['dtype']} | {muestras} |")
        L.append("")
    return "\n".join(L)


def main():
    ap = argparse.ArgumentParser(description="Cataloga Excel/CSV gigantes a .md navegables")
    ap.add_argument("--root", required=True)
    ap.add_argument("--salida", required=True)
    ap.add_argument("--muestras", type=int, default=5, help="filas de muestra por hoja")
    ap.add_argument("--max-cols", type=int, default=200, help="tope de columnas a catalogar por hoja")
    args = ap.parse_args()

    root = os.path.abspath(args.root)
    salida = os.path.abspath(args.salida)
    os.makedirs(salida, exist_ok=True)

    if not os.path.isdir(root):
        log(f"NO EXISTE root: {root}"); sys.exit(1)

    catalogo = {"root": root, "generado": datetime.now().isoformat(), "archivos": []}
    indice = ["# Índice de fuentes históricas", "", f"Root: `{root}`  ·  generado {catalogo['generado']}", "",
              "| Archivo | MB | Hojas | Filas (máx hoja) |", "|---|---|---|---|"]

    archivos = []
    for dirpath, _, files in os.walk(root):
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in EXT_EXCEL or ext in EXT_CSV:
                archivos.append(os.path.join(dirpath, f))
    archivos.sort()
    log(f"{len(archivos)} archivos a catalogar")

    for k, path in enumerate(archivos, 1):
        rel = os.path.relpath(path, root).replace("\\", "/")
        ext = os.path.splitext(path)[1].lower()
        mb = round(os.path.getsize(path) / (1024 * 1024), 2)
        log(f"[{k}/{len(archivos)}] {rel} ({mb} MB)")
        try:
            hojas = procesar_excel(path, args.muestras, args.max_cols) if ext in EXT_EXCEL \
                else procesar_csv(path, args.muestras, args.max_cols)
        except Exception as e:
            hojas = [{"nombre": "(archivo)", "error": str(e)}]
            log(f"   ⚠️ error: {e}")

        meta = {"rel": rel, "ext": ext, "tamano_mb": mb, "hojas": hojas}
        catalogo["archivos"].append(meta)

        # .md espejo por archivo
        md_path = os.path.join(salida, rel + ".md")
        os.makedirs(os.path.dirname(md_path), exist_ok=True)
        with open(md_path, "w", encoding="utf-8") as fh:
            fh.write(md_de_archivo(rel, meta))

        max_filas = max((h.get("filas_datos", 0) for h in hojas), default=0)
        indice.append(f"| [{rel}]({rel}.md) | {mb} | {len(hojas)} | {max_filas} |")

    with open(os.path.join(salida, "INDICE.md"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(indice) + "\n")
    with open(os.path.join(salida, "catalogo.json"), "w", encoding="utf-8") as fh:
        json.dump(catalogo, fh, ensure_ascii=False, indent=1, default=str)

    log(f"OK → {salida}  (INDICE.md + catalogo.json + {len(archivos)} .md)")


if __name__ == "__main__":
    main()
