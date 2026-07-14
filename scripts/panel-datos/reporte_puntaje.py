# -*- coding: utf-8 -*-
"""
reporte_puntaje.py — Ranking de estudiantes por puntaje compuesto ("calidad de estudiante").

Lee la vista `v_puntaje_estudiante` de Supabase (JC, cohorte actual) y produce el ranking.
El puntaje pondera señales convertidas a PERCENTIL dentro de la cohorte:

    ingresos Emoflow 60%  ·  avance Q10 40%  ·  asistencia Zoom 0%

REGLA DE NEGOCIO (2026-07-14): **Emoflow es el criterio mayor y es obligatorio** — el estudiante
sin ingresos registrados NO entra al ranking. Los pesos se ajustan con --peso-ingresos /
--peso-avance / --peso-asistencia.

⚠ Por qué percentiles: los valores crudos mienten. `avance_q10` promedia 92.8 con sd 6.7 (casi
  no discrimina), así que con un peso nominal alto aportaba MENOS al ranking que las otras
  señales. Además, renormalizando sobre crudos, a quien le faltaba una señal el avance (~93) le
  apuntalaba el puntaje: faltar dato PREMIABA. Con percentiles las señales quedan uniformes en
  [0,100] → los pesos significan lo que dicen y quitar una señal no sesga.

⚠ La asistencia arranca con peso 0 porque NO es una señal madura: cubre 408/777, viene de UN solo
  curso ("Desarrollo Web - GIT, HTML y CSS") y lleva ~11 días de captura → 1.4 sesiones por
  persona (solo 4 estudiantes con >=3). Un promedio sobre 1 sesión es ruido. Cuando acumule
  sesiones y cubra más cursos, se le da peso con --peso-asistencia.

⚠ PRIVACIDAD: la salida lleva nombre y correo (PII) → va a tools/ (gitignoreado). Nunca a GitHub.

Uso:
    python reporte_puntaje.py                 # top 25 + CSV (Emoflow 60% + avance 40%)
    python reporte_puntaje.py --top 50
    python reporte_puntaje.py --ciudad BOG
    python reporte_puntaje.py --ciudad BOG --limite 100 --excel "%USERPROFILE%\\Downloads\\100 mejores de bogota.xlsx"
    python reporte_puntaje.py --peso-ingresos 0.8 --peso-avance 0.2     # Emoflow aún más dominante
    python reporte_puntaje.py --peso-asistencia 0.2                     # cuando la asistencia madure

Fundación ROFÉ | Jóvenes creaTIvos
"""

import argparse
import csv
import io
import json
import os
import sys
import urllib.error
import urllib.request
from datetime import datetime

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT     = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RUTA_ENV          = os.path.join(PROYECTO_ROOT, ".env.local")
USER_AGENT        = "panel-datos-etl/1.0"

COLUMNAS = ["puesto", "nombre", "email", "grupo_ciudad", "puntaje", "senales",
            "avance_q10", "pct_avance", "asistencia", "sesiones", "pct_asistencia",
            "ingresos_emoflow", "pct_ingresos", "ultimo_ingreso"]


def cargar_env_local() -> None:
    if not os.path.isfile(RUTA_ENV):
        return
    with open(RUTA_ENV, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())


def get_todo(url: str, key: str, ruta: str, page: int = 1000) -> list:
    filas, offset = [], 0
    base = url.rstrip("/") + "/rest/v1"
    sep = "&" if "?" in ruta else "?"
    while True:
        req = urllib.request.Request(
            f"{base}{ruta}{sep}limit={page}&offset={offset}",
            headers={"apikey": key, "Authorization": f"Bearer {key}",
                     "User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                lote = json.loads(resp.read())
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code}: {e.read().decode(errors='replace')[:300]}") from None
        filas.extend(lote)
        if len(lote) < page:
            return filas
        offset += page


def main() -> int:
    ap = argparse.ArgumentParser(description="Ranking de estudiantes por puntaje compuesto")
    ap.add_argument("--top", type=int, default=25, help="cuántos mostrar en consola (default 25)")
    ap.add_argument("--ciudad", help="filtrar por grupo_ciudad (BAQ, BOG, CAL, ...)")
    ap.add_argument("--limite", type=int, help="quedarse solo con los N mejores (ej. 100)")
    ap.add_argument("--excel", help="ruta del .xlsx a generar (además del CSV)")
    ap.add_argument("--peso-ingresos", type=float, default=0.60,
                    help="peso del percentil de ingresos Emoflow — criterio MAYOR (default 0.60)")
    ap.add_argument("--peso-avance", type=float, default=0.40,
                    help="peso del percentil de avance Q10 (default 0.40)")
    ap.add_argument("--peso-asistencia", type=float, default=0.0,
                    help="peso del percentil de asistencia Zoom (default 0 — señal aún inmadura: "
                         "un solo curso y ~1.4 sesiones por persona)")
    args = ap.parse_args()

    cargar_env_local()
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: falta SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY en .env.local")
        return 1

    filas = get_todo(url, key, "/v_puntaje_estudiante?select=*")

    # La ciudad sale de la BD de monitorías (participants.grupo_ciudad) y tiene huecos.
    # El `Area` de Emoflow sirve de respaldo — recupera a los que la BD no ubicó.
    emo_ciudad = {e["participant_id"]: e["grupo_ciudad"]
                  for e in get_todo(url, key,
                                    "/emoflow_ingresos?select=participant_id,grupo_ciudad")
                  if e.get("participant_id") and e.get("grupo_ciudad")}
    for f in filas:
        if not f.get("grupo_ciudad"):
            f["grupo_ciudad"] = emo_ciudad.get(f["id"])

    if args.ciudad:
        filas = [f for f in filas if (f.get("grupo_ciudad") or "") == args.ciudad.upper()]

    # REGLA DE NEGOCIO: Emoflow es el criterio mayor y es OBLIGATORIO.
    # Sin ingresos registrados el estudiante no entra al ranking (no "cuenta").
    antes = len(filas)
    filas = [f for f in filas if f.get("pct_ingresos") is not None]
    sin_emoflow = antes - len(filas)
    if sin_emoflow:
        print(f"Excluidos {sin_emoflow} sin dato de Emoflow (regla: sin Emoflow no cuenta)")

    # Puntaje = promedio ponderado de los PERCENTILES de cada señal, renormalizado sobre las
    # presentes. Emoflow manda; el avance acompaña; la asistencia solo si se le da peso
    # explícito (default 0: hoy es una señal inmadura — 1 solo curso, ~1.4 sesiones/persona).
    pesos = {"pct_ingresos": args.peso_ingresos,
             "pct_avance": args.peso_avance,
             "pct_asistencia": args.peso_asistencia}
    for f in filas:
        num = den = 0.0
        for clave, w in pesos.items():
            v = f.get(clave)
            if w > 0 and v is not None:
                num += float(v) * w
                den += w
        f["puntaje"] = round(num / den, 1) if den else None

    filas = [f for f in filas if f.get("puntaje") is not None]
    filas.sort(key=lambda f: f["puntaje"], reverse=True)
    if args.limite:
        filas = filas[:args.limite]
    for i, f in enumerate(filas, 1):
        f["puesto"] = i

    partes = [f"ingresos Emoflow {args.peso_ingresos:.0%}", f"avance Q10 {args.peso_avance:.0%}"]
    if args.peso_asistencia > 0:
        partes.append(f"asistencia {args.peso_asistencia:.0%}")
    etiqueta = " + ".join(partes) + " (sobre percentiles; sin Emoflow no cuenta)"
    print(f"\nPuntaje: {etiqueta}")
    print(f"Estudiantes rankeados: {len(filas)}\n")

    anchos = (4, 38, 6, 8, 9, 9, 9)
    print(f"{'#':<4} {'Nombre':<38} {'Ciud':<6} {'Puntaje':>8} {'Avance':>9} {'Ingresos':>9} {'Asist':>9}")
    print("-" * 92)
    for f in filas[:args.top]:
        asis = f"{f['asistencia']} ({f['sesiones']}s)" if f.get("asistencia") is not None else "—"
        print(f"{f['puesto']:<4} {(f['nombre'] or '')[:38]:<38} {(f['grupo_ciudad'] or '—'):<6} "
              f"{float(f['puntaje']):>8.1f} {str(f.get('avance_q10') or '—'):>9} "
              f"{str(f.get('ingresos_emoflow') or '—'):>9} {asis:>9}")

    salida = os.path.join(PROYECTO_ROOT, "tools",
                          f"puntaje_estudiantes_{datetime.now():%Y%m%d}.csv")
    os.makedirs(os.path.dirname(salida), exist_ok=True)
    with open(salida, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNAS, extrasaction="ignore")
        w.writeheader()
        w.writerows(filas)

    print(f"\nCSV completo ({len(filas)} filas, PII) → {os.path.relpath(salida, PROYECTO_ROOT)}")

    if args.excel:
        escribir_excel(filas, os.path.expandvars(args.excel), etiqueta, args.ciudad)

    print(f"RESUMEN: rankeados={len(filas)} sin_emoflow_excluidos={sin_emoflow} "
          f"pesos=ingresos{args.peso_ingresos}/avance{args.peso_avance}/"
          f"asistencia{args.peso_asistencia} estado=exito")
    return 0


ENCABEZADOS = [
    ("puesto", "#", 6), ("nombre", "Nombre", 34), ("email", "Correo", 34),
    ("grupo_ciudad", "Ciudad", 9), ("puntaje", "Puntaje", 10),
    ("avance_q10", "Avance Q10 %", 13), ("pct_avance", "Percentil avance", 16),
    ("ingresos_emoflow", "Ingresos Emoflow", 17), ("pct_ingresos", "Percentil ingresos", 18),
    ("asistencia", "Asistencia %", 13), ("sesiones", "Sesiones", 10),
    ("senales", "Señales", 9), ("ultimo_ingreso", "Último ingreso", 15),
]


def escribir_excel(filas: list, ruta: str, etiqueta: str, ciudad: str | None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (ciudad or "Ranking")[:31]

    azul   = PatternFill("solid", fgColor="1F4E79")
    gris   = PatternFill("solid", fgColor="F2F2F2")
    blanco = Font(color="FFFFFF", bold=True)

    titulo = f"Top {len(filas)} — {ciudad or 'Jóvenes creaTIvos'}"
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Puntaje: {etiqueta}"])
    ws.append([f"Generado {datetime.now():%Y-%m-%d %H:%M} · Fundación ROFÉ / Jóvenes creaTIvos"])
    ws.append([])

    fila_head = 5
    for i, (_, titulo_col, ancho) in enumerate(ENCABEZADOS, 1):
        c = ws.cell(row=fila_head, column=i, value=titulo_col)
        c.fill, c.font = azul, blanco
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho

    for n, f in enumerate(filas):
        for i, (clave, _, _) in enumerate(ENCABEZADOS, 1):
            v = f.get(clave)
            if clave in ("puntaje", "avance_q10", "pct_avance", "pct_ingresos",
                         "asistencia") and v is not None:
                v = float(v)
            c = ws.cell(row=fila_head + 1 + n, column=i, value=v if v is not None else "—")
            if n % 2 == 1:
                c.fill = gris
            if clave == "puntaje":
                c.font = Font(bold=True)
            if clave in ("puntaje", "avance_q10", "pct_avance", "pct_ingresos", "asistencia"):
                c.number_format = "0.0"

    ws.freeze_panes = ws.cell(row=fila_head + 1, column=1)
    ws.auto_filter.ref = (f"A{fila_head}:"
                          f"{get_column_letter(len(ENCABEZADOS))}{fila_head + len(filas)}")

    os.makedirs(os.path.dirname(ruta) or ".", exist_ok=True)
    wb.save(ruta)
    print(f"Excel ({len(filas)} filas, PII) → {ruta}")


if __name__ == "__main__":
    sys.exit(main())
