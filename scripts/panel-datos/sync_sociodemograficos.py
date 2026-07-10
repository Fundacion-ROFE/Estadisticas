# -*- coding: utf-8 -*-
"""
sync_sociodemograficos.py — BD Seguimiento de Monitorias (xlsx) → Supabase participants.

Extrae de la pestaña `Seguimiento` (hub, headers fila 1): ID(c7), Grupo(c2),
Fecha Nacimiento(c11), Edad(c12), Ciudad(c13), Género(c16); y de `Diagnostico`
(headers fila 1): Número de documento(c3) + situación de emprendimiento(c32,
4 categorías). Cruza por cédula normalizada y actualiza SOLO participantes que
ya existen en Supabase (los de la BD que no están en Q10 se reportan, no se crean).

tiene_emprendimiento = (situacion == en_marcha). Al final recomputa agregados.

⚠ PRIVACIDAD: PII en memoria y en tools/ solamente. La RUTA_BD apunta al export
xlsx en Downloads — actualizarla cuando cambie la versión del archivo (la fuente
viva es un Google Sheet; ver docs/procesos/bd-seguimiento-monitorias.md).

Uso:
    python sync_sociodemograficos.py [--dry-run]
Consola (parseable por n8n):
    RESUMEN: actualizados=N sin_match_supabase=X sin_datos=Y estado=exito

Fundación ROFÉ | Jóvenes creaTIvos
"""

import argparse
import io
import json
import os
import re
import sys
import urllib.error
import urllib.request
from collections import Counter
from datetime import date, datetime

try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT     = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RUTA_ENV          = os.path.join(PROYECTO_ROOT, ".env.local")
RUTA_BD           = r"C:\Users\EstudiantesJC\Downloads\BD Seguimiento de Monitorias - JC2026 (1).xlsx"
RUTA_REPORTE      = os.path.join(PROYECTO_ROOT, "tools",
                                 f"sociodemograficos_report_{datetime.now():%Y%m%d}.json")

USER_AGENT = "panel-datos-etl/1.0"  # Supabase rechaza secrets con UA de navegador
LOTE       = 500

# Diagnostico c32 → enum emprendimiento_situacion (match por prefijo)
MAPA_EMPRENDIMIENTO = [
    ("sí, tengo un emprendimiento en marcha", "en_marcha"),
    ("tengo una idea de negocio",             "idea"),
    ("me llama la atención emprender",        "interesado"),
    ("no me interesa emprender",              "no_interesado"),
]


def log(msg: str) -> None:
    print(f"[sync-sociodem] {msg}", flush=True)


def norm_id(valor) -> str:
    """Cédula a solo dígitos. Gotcha openpyxl: las celdas numéricas llegan como
    float (1041774123.0) — str() directo mete '.0' y el strip de no-dígitos lo
    convertía en un CERO EXTRA al final (10417741230 ≠ 1041774123)."""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return re.sub(r"\D", "", str(valor))


def cargar_env_local() -> None:
    if not os.path.isfile(RUTA_ENV):
        return
    with open(RUTA_ENV, encoding="utf-8") as f:
        for linea in f:
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            k, v = linea.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())


class Supa:
    def __init__(self, url: str, key: str):
        self.base = url.rstrip("/") + "/rest/v1"
        self.key = key

    def _req(self, metodo: str, ruta: str, cuerpo=None, prefer: str = ""):
        headers = {"apikey": self.key, "Authorization": f"Bearer {self.key}",
                   "Content-Type": "application/json", "User-Agent": USER_AGENT}
        if prefer:
            headers["Prefer"] = prefer
        req = urllib.request.Request(self.base + ruta, method=metodo, headers=headers,
                                     data=json.dumps(cuerpo).encode() if cuerpo is not None else None)
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                datos = resp.read()
                return resp.status, json.loads(datos) if datos else None
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} en {metodo} {ruta}: "
                               f"{e.read().decode(errors='replace')[:500]}") from None

    def get_todo(self, ruta: str, page: int = 1000) -> list:
        filas, offset = [], 0
        sep = "&" if "?" in ruta else "?"
        while True:
            _, lote = self._req("GET", f"{ruta}{sep}limit={page}&offset={offset}")
            filas.extend(lote or [])
            if not lote or len(lote) < page:
                return filas
            offset += page


def extraer_bd() -> dict:
    """{cedula: {genero, fecha_nacimiento, edad, ciudad, grupo_ciudad,
                 situacion_emprendimiento, tiene_emprendimiento}}"""
    log(f"Leyendo {os.path.basename(RUTA_BD)}...")
    wb = load_workbook(RUTA_BD, read_only=True, data_only=True)

    datos: dict[str, dict] = {}
    ws = wb["Seguimiento"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 16:
            continue
        ced = norm_id(r[6])          # c7 ID
        if not ced:
            continue
        fnac = r[10]                 # c11 Fecha Nacimiento (datetime)
        edad = r[11]                 # c12 Edad (numérica)
        datos[ced] = {
            "grupo_ciudad": str(r[1]).strip() if r[1] else None,          # c2 Grupo
            "fecha_nacimiento": fnac.date().isoformat()
                                if isinstance(fnac, datetime) else
                                (fnac.isoformat() if isinstance(fnac, date) else None),
            "edad": int(edad) if isinstance(edad, (int, float)) else None,
            "ciudad": str(r[12]).strip() if r[12] else None,              # c13 Ciudad
            "genero": str(r[15]).strip() if r[15] else None,              # c16 Género
        }
    log(f"Seguimiento: {len(datos)} cédulas con datos")

    ws2 = wb["Diagnostico"]
    n_emp = 0
    for r in ws2.iter_rows(min_row=2, values_only=True):
        if not r or len(r) < 32:
            continue
        ced = norm_id(r[2])          # c3 Número de documento
        if not ced or not r[31]:     # c32 situación emprendimiento
            continue
        texto = str(r[31]).strip().lower()
        situacion = next((v for pref, v in MAPA_EMPRENDIMIENTO if texto.startswith(pref)), None)
        if situacion is None:
            continue
        fila = datos.setdefault(ced, {})
        fila["situacion_emprendimiento"] = situacion
        fila["tiene_emprendimiento"] = situacion == "en_marcha"
        n_emp += 1
    wb.close()
    log(f"Diagnostico: {n_emp} respuestas de emprendimiento mapeadas")
    return datos


def main() -> int:
    ap = argparse.ArgumentParser(description="BD monitorias → Supabase participants")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    cargar_env_local()
    url, key = os.environ.get("SUPABASE_URL", ""), os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        log("ERROR: faltan SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")
        return 1
    if not os.path.isfile(RUTA_BD):
        log(f"ERROR: no existe {RUTA_BD} — actualizar RUTA_BD a la versión vigente")
        return 1

    datos = extraer_bd()
    supa = Supa(url, key)
    # q10_id → nombre actual: el eco del nombre satisface el NOT NULL del INSERT
    # propuesto en el upsert (Postgres valida constraints ANTES de resolver el conflicto)
    existentes = {r["q10_id"]: r["nombre"]
                  for r in supa.get_todo("/participants?select=q10_id,nombre")}
    log(f"Participantes en Supabase: {len(existentes)}")

    filas, sin_match, sin_datos, stats = [], [], 0, Counter()
    for ced, d in sorted(datos.items()):
        limpio = {k: v for k, v in d.items() if v is not None}
        if not limpio:
            sin_datos += 1
            continue
        if ced not in existentes:
            sin_match.append(ced)   # en BD monitorias pero no en Q10 (retirados, típicamente)
            continue
        filas.append({"q10_id": ced, "nombre": existentes[ced], **limpio,
                      "updated_at": datetime.now().isoformat(timespec="seconds")})
        for k in limpio:
            stats[k] += 1

    log(f"A actualizar: {len(filas)} · sin match en Supabase: {len(sin_match)} · sin datos: {sin_datos}")
    log("Campos: " + ", ".join(f"{k}={v}" for k, v in stats.most_common()))

    if args.dry_run:
        print(f"RESUMEN: actualizados=0 sin_match_supabase={len(sin_match)} "
              f"sin_datos={sin_datos} estado=dry_run")
        return 0

    # upsert por q10_id: en filas existentes PostgREST solo toca las columnas provistas.
    # Gotcha PGRST102: el bulk exige claves idénticas en todas las filas del batch →
    # agrupar por conjunto de claves (así "vacío nunca sobreescribe" se mantiene:
    # jamás mandamos null explícito por un dato que la BD no trae).
    grupos: dict[frozenset, list] = {}
    for fila in filas:
        grupos.setdefault(frozenset(fila), []).append(fila)
    for claves, grupo in grupos.items():
        for i in range(0, len(grupo), LOTE):
            supa._req("POST", "/participants?on_conflict=q10_id", grupo[i:i + LOTE],
                      prefer="resolution=merge-duplicates,return=minimal")
    log(f"Actualizados: {len(filas)} (en {len(grupos)} grupos de columnas)")

    _, agg = supa._req("POST", "/rpc/recompute_aggregates", {})
    log(f"Agregados recomputados: {agg}")

    with open(RUTA_REPORTE, "w", encoding="utf-8") as f:
        json.dump({"generado": datetime.now().isoformat(timespec="seconds"),
                   "actualizados": len(filas), "campos": dict(stats),
                   "sin_match_supabase": sorted(sin_match), "sin_datos": sin_datos},
                  f, ensure_ascii=False, indent=1)
    log(f"Reporte → {RUTA_REPORTE}")

    print(f"RESUMEN: actualizados={len(filas)} sin_match_supabase={len(sin_match)} "
          f"sin_datos={sin_datos} estado=exito")
    return 0


if __name__ == "__main__":
    sys.exit(main())
