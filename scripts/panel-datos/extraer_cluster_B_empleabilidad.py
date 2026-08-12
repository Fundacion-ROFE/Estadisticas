# -*- coding: utf-8 -*-
"""
extraer_cluster_B_empleabilidad.py
==================================
Clúster B — enriquecimiento EMPLEABILIDAD (Jóvenes creaTIvos).

Extrae la dimensión de EMPLEABILIDAD que HOY NO existe en Supabase (el equipo la ve
en Power BI y nosotros no): la intención de vincularse laboralmente, si la persona
figura en un listado de empleabilidad de su cohorte, y si presentó el proyecto final.
Cada valor se ancla a la persona conocida (canon) con el Matcher compartido
(prioridad cédula > correo > nombre).

FUENTES (raíz COMPLETE-ORDEN-INFORMATION):
  - 6-2024/Copia de empleabilidad 22-23.xlsx  (12 hojas ciudad-año; -22→2022, -23→2023)
  - 2-2020/Empleabilidad JC2020.xlsx                     (2020)
  - 2-2020/Empleabilidad JC2020 Enero 21.xlsx            (2020)
  - 3-2021/Empleabilidad JC2021.xlsx                     (2021)
  - 3-2021/Empleabilidad.xlsx                            (2021, duplicado de la anterior)
  - 4-2022/Listado Empleabilidad 2022.xlsx               (2022)
  - 5-2023/Listado Empleabilidad 2023.xlsx               (2023)

OJO — trampa de datos (documentada en el digest y verificada en vivo):
  En las hojas «-23» del archivo 22-23 las columnas están CORRIDAS respecto a las «-22»:
  el encabezado "Me gustaría vincularme..." termina con la INSTITUCIÓN, y la respuesta
  de intención ("SI, definitivamente quiero trabajar en tecnología...") cae en la
  columna "EMPLEABILIDAD". Por eso NO se confía en el nombre del encabezado para la
  intención: se detecta por CONTENIDO de la celda. La identidad (# DOCUMENTO, CORREO,
  NOMBRE) sí es estable y se toma por encabezado.

CAMPOS OBJETIVO → cómo se obtienen:
  intencion_empleo     ← celda con la respuesta canónica de intención (content-scan)
  aplica_empleabilidad ← "si" para toda persona que figure en un listado de empleabilidad
  proyecto_final       ← columna PROYECTO/PROYECTO FINAL con valor SI/OK (extra útil)
  vinculado_laboralmente / empresa / cargo → NO existen en ninguna de estas fuentes
                          (no se inventan; se reporta su ausencia en el RESUMEN).

SALIDAS (PII → SOLO tools/, gitignoreado; NUNCA a docs/ ni a Supabase):
  - tools/enriquecimiento/B_empleabilidad.json
      lista de {canon, cedula_fuente, campo, valor, fuente_archivo, hoja, anio, metodo_match}
  - tools/enriquecimiento/B_RESUMEN.md

Solo lectura sobre los xlsx. Idempotente / reejecutable.
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from enriquecimiento_helper import Matcher, cargar_hoja, norm_cedula  # noqa: E402

# --------------------------------------------------------------------------- #
# Configuración
# --------------------------------------------------------------------------- #
RAIZ = Path(r"C:\Users\EstudiantesJC\Downloads\COMPLETE-ORDEN-INFORMATION")
OUT_DIR = Path(r"C:\Users\EstudiantesJC\downloads\admin-usable\tools\enriquecimiento")

# (ruta_relativa, hoja|None=todas, anio|None=derivar de la hoja)
FUENTES = [
    ("6-2024/Copia de empleabilidad 22-23.xlsx", None, None),  # 12 hojas, año por sufijo
    ("2-2020/Empleabilidad JC2020.xlsx", "Empleabilidad", 2020),
    ("2-2020/Empleabilidad JC2020 Enero 21.xlsx", "Empleabilidad", 2020),
    ("3-2021/Empleabilidad JC2021.xlsx", "Empleabilidad", 2021),
    ("3-2021/Empleabilidad.xlsx", "Empleabilidad", 2021),
    ("4-2022/Listado Empleabilidad 2022.xlsx", "Empleabilidad", 2022),
    ("5-2023/Listado Empleabilidad 2023.xlsx", "Empleabilidad", 2023),
]

# Respuesta canónica de intención de empleo (content-scan, robusto al corrimiento de
# columnas). Excluye textos de "proyecto final" que solo mencionan "tal vez" de pasada.
INTENCION_RX = re.compile(
    r"(trabajar en tecnolog|no he decidido|^\s*tal\s*vez\b|"
    r"^\s*s[ií],\s*definitivamente|^\s*no,)",
    re.I,
)

BASURA = {"", "n/a", "na", "null", "none", "#n/a", "#ref!", "-", "."}
SI_OK = {"si", "sí", "ok", "x"}


def limpiar(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in BASURA else s


def find_col(cols, incluir, excluir=None):
    """Primer encabezado que matchee `incluir` (regex) y no matchee `excluir`."""
    inc = re.compile(incluir, re.I)
    exc = re.compile(excluir, re.I) if excluir else None
    for c in cols:
        if inc.search(c) and not (exc and exc.search(c)):
            return c
    return None


def anio_de_hoja(hoja: str):
    m = re.search(r"-\s*(\d{2})\s*$", hoja.strip())
    if m:
        return 2000 + int(m.group(1))
    return None


def detectar_intencion(fila, cols):
    """Devuelve el texto de la respuesta de intención (o "") buscando por contenido."""
    for c in cols:
        v = fila.get(c)
        if isinstance(v, str) and INTENCION_RX.search(v):
            return v.strip()
    return ""


def detectar_proyecto(fila, cols):
    """'si' si alguna columna PROYECTO tiene un valor SI/OK."""
    for c in cols:
        if re.search(r"proyecto", c, re.I):
            v = limpiar(fila.get(c)).lower()
            if v in SI_OK:
                return "si"
    return ""


def main():
    m = Matcher.desde_archivo()

    payload = []
    faltantes = []
    # métricas
    metodo_por_persona = {}                       # canon -> primer método
    canon_por_anio = defaultdict(set)             # anio -> {canon}
    canon_intencion = set()                        # canon con intención
    intencion_valores = Counter()                  # texto de intención -> n
    cobertura_campo = defaultdict(set)             # campo -> {canon}
    ejemplos = {}
    filas_por_fuente = Counter()
    # dedup a nivel (canon, campo, fuente_archivo, hoja, anio)
    emitidos = set()

    for rel, hoja_fija, anio_fijo in FUENTES:
        ruta = RAIZ / rel
        if not ruta.exists():
            faltantes.append(rel)
            print(f"[B] AVISO: no existe la fuente: {rel}")
            continue

        from openpyxl import load_workbook
        wb = load_workbook(str(ruta), read_only=True, data_only=True)
        hojas = [hoja_fija] if hoja_fija else list(wb.sheetnames)
        wb.close()

        for hoja in hojas:
            cols, filas = cargar_hoja(str(ruta), sheet=hoja)
            anio = anio_fijo if anio_fijo else anio_de_hoja(hoja)

            col_ced = find_col(cols, r"#?\s*documento|n[uú]mero de identif|c[eé]dula",
                               excluir=r"tipo")
            col_email = find_col(cols, r"correo|e-?mail")
            col_nombre = find_col(cols, r"nombre")

            for fila in filas:
                ced_fuente = norm_cedula(fila.get(col_ced)) if col_ced else ""
                canon, metodo = m.match(
                    cedula=fila.get(col_ced) if col_ced else None,
                    email=fila.get(col_email) if col_email else None,
                    nombre=fila.get(col_nombre) if col_nombre else None,
                )
                if not canon:
                    continue
                filas_por_fuente[rel] += 1
                metodo_por_persona.setdefault(canon, metodo)
                if anio:
                    canon_por_anio[anio].add(canon)

                registros = []  # (campo, valor)
                # 1) aparecer en el listado => aplica a empleabilidad
                registros.append(("aplica_empleabilidad", "si"))
                # 2) intención (por contenido)
                intent = detectar_intencion(fila, cols)
                if intent:
                    registros.append(("intencion_empleo", intent))
                    canon_intencion.add(canon)
                    intencion_valores[intent] += 1
                # 3) proyecto final (extra útil)
                proy = detectar_proyecto(fila, cols)
                if proy:
                    registros.append(("proyecto_final", proy))

                for campo, valor in registros:
                    clave = (canon, campo, rel, hoja, anio)
                    if clave in emitidos:
                        continue
                    emitidos.add(clave)
                    cobertura_campo[campo].add(canon)
                    ejemplos.setdefault(campo, valor)
                    payload.append({
                        "canon": canon,
                        "cedula_fuente": ced_fuente,
                        "campo": campo,
                        "valor": valor,
                        "fuente_archivo": rel,
                        "hoja": hoja.strip(),
                        "anio": anio,
                        "metodo_match": metodo,
                    })

    # ---- Escritura ----
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "B_empleabilidad.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    personas = set(metodo_por_persona)
    n_personas = len(personas)
    metodos = Counter(metodo_por_persona.values())

    escribir_resumen(payload, n_personas, metodos, canon_por_anio, cobertura_campo,
                     canon_intencion, intencion_valores, ejemplos, filas_por_fuente,
                     faltantes)

    # ---- Consola ----
    print("=== CLÚSTER B — EMPLEABILIDAD ===")
    print(f"Personas (canon) con match: {n_personas}")
    print(f"Registros en payload: {len(payload)}")
    print("Match por método (personas):")
    for met, n in metodos.most_common():
        pct = 100 * n / n_personas if n_personas else 0
        print(f"   {met:8} {n:5}  ({pct:.1f}%)")
    print("Cobertura por campo (personas con valor):")
    for campo in ("aplica_empleabilidad", "intencion_empleo", "proyecto_final"):
        print(f"   {campo:22} {len(cobertura_campo.get(campo, set())):5}")
    print("Personas con empleabilidad por año:")
    for anio in sorted(k for k in canon_por_anio if k):
        print(f"   {anio}: {len(canon_por_anio[anio])}")
    print(f"\nSalidas en: {OUT_DIR}")


def escribir_resumen(payload, n_personas, metodos, canon_por_anio, cobertura_campo,
                     canon_intencion, intencion_valores, ejemplos, filas_por_fuente,
                     faltantes):
    L = []
    L.append("# Clúster B — EMPLEABILIDAD (Jóvenes creaTIvos)\n")
    L.append("Dimensión de empleabilidad que hoy NO está en Supabase (el equipo la ve "
             "en Power BI). Match por cédula/correo/nombre contra el roster de identidad.\n")

    L.append("## Resultado global\n")
    L.append(f"- Personas (canon) con match: **{n_personas}**")
    L.append(f"- Registros (canon, campo, fuente) en el payload: **{len(payload)}**")
    L.append(f"- Personas con respuesta de **intención** de empleo: "
             f"**{len(canon_intencion)}**\n")
    if faltantes:
        L.append(f"> AVISO: fuentes no encontradas: {faltantes}\n")

    L.append("## Match por método (personas)\n")
    L.append("| Método | Personas | % |")
    L.append("|---|---|---|")
    for met, n in metodos.most_common():
        pct = 100 * n / n_personas if n_personas else 0
        L.append(f"| {met} | {n} | {pct:.1f}% |")
    L.append("")

    L.append("## Personas con empleabilidad por año\n")
    L.append("Año = sufijo de la hoja (-22/-23) o el año del archivo. Una persona puede "
             "figurar en más de un año (p. ej. lista 2022 y seguimiento 2023).\n")
    L.append("| Año | Personas |")
    L.append("|---|---|")
    for anio in sorted(k for k in canon_por_anio if k):
        L.append(f"| {anio} | {len(canon_por_anio[anio])} |")
    L.append("")

    L.append("## Cobertura por campo\n")
    L.append("| Campo | Personas con valor | Nota |")
    L.append("|---|---|---|")
    notas = {
        "aplica_empleabilidad": "figura en un listado de empleabilidad",
        "intencion_empleo": "respuesta textual de intención (solo 2020 y 22-23)",
        "proyecto_final": "presentó proyecto final (SI/OK)",
    }
    for campo in ("aplica_empleabilidad", "intencion_empleo", "proyecto_final"):
        L.append(f"| {campo} | {len(cobertura_campo.get(campo, set()))} "
                 f"| {notas[campo]} |")
    L.append("")
    L.append("> `vinculado_laboralmente`, `empresa` y `cargo` **no existen** en ninguna "
             "de estas fuentes (son listados de intención/candidatura, no de "
             "colocación laboral). No se emitieron ni se inventaron.\n")

    L.append("## Distribución de la intención de empleo\n")
    L.append("| Respuesta | Registros |")
    L.append("|---|---|")
    for v, n in intencion_valores.most_common():
        vv = v.replace("\n", " ").replace("|", "/")
        if len(vv) > 70:
            vv = vv[:67] + "..."
        L.append(f"| {vv} | {n} |")
    L.append("")

    L.append("## Registros por fuente (filas con match)\n")
    L.append("| Fuente | Filas con match |")
    L.append("|---|---|")
    for f, n in filas_por_fuente.most_common():
        L.append(f"| `{f}` | {n} |")
    L.append("")

    L.append("## Ejemplos de valor por campo\n")
    L.append("| Campo | Ejemplo |")
    L.append("|---|---|")
    for campo in ("aplica_empleabilidad", "intencion_empleo", "proyecto_final"):
        ej = str(ejemplos.get(campo, "—")).replace("\n", " ").replace("|", "/")
        if len(ej) > 80:
            ej = ej[:77] + "..."
        L.append(f"| {campo} | {ej} |")
    L.append("")

    (OUT_DIR / "B_RESUMEN.md").write_text("\n".join(L), encoding="utf-8")


if __name__ == "__main__":
    main()
