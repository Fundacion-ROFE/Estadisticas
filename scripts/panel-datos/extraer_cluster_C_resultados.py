# -*- coding: utf-8 -*-
"""
extraer_cluster_C_resultados.py — Clúster C (RESULTADOS / PROYECTO FINAL / PUNTAJES)
de la etapa final de enriquecimiento (Fundación ROFÉ / Jóvenes creaTIvos).

Extrae de los archivos históricos de fin de programa:
  - presento_proyecto (sí/no)   ← listas de presentación de proyecto final 2021-2024
  - nota_final                  ← hoja NOTA AVANZADOS/BÁSICOS 2024 (columna TOTAL)
  - mesa / categoria            ← Expoferia por mesa 2023; GRUPO donde exista
  - puntaje_seleccion_fase2     ← PUNTUACION FINALISTAS 2019 (OJO: es puntaje de
                                  SELECCIÓN de finalistas, NO del proyecto final; se
                                  incluye por completitud pero con campo distinto)

NO existen en las fuentes columnas separadas de "puntaje_jurado"/"puntaje_tecnico"
ni "nombre_proyecto": las listas son cronogramas de presentación, no rúbricas de
jurado. Ver C_RESUMEN.md para el detalle honesto por fuente.

Identidad: se ancla cada fila a una persona conocida con enriquecimiento_helper.Matcher
(prioridad cédula > correo > nombre). Muchas de estas fuentes identifican por NOMBRE,
así que se registra `nombre_crudo` en cada registro para poder auditar homónimos.

PII: salida SOLO en tools/ (gitignoreado). No toca Supabase ni modifica las fuentes.
Idempotente: reescribe el payload en cada corrida.

Uso:
    python extraer_cluster_C_resultados.py
"""

import io
import json
import os
import re
import sys

sys.path.insert(0, r"C:\Users\EstudiantesJC\downloads\admin-usable\scripts\panel-datos")
from enriquecimiento_helper import Matcher, norm_cedula, norm_email, norm_nombre  # noqa: E402

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

RAIZ = r"C:\Users\EstudiantesJC\Downloads\COMPLETE-ORDEN-INFORMATION"
SALIDA_DIR = r"C:\Users\EstudiantesJC\downloads\admin-usable\tools\enriquecimiento"
PAYLOAD = os.path.join(SALIDA_DIR, "C_resultados.json")
RESUMEN = os.path.join(SALIDA_DIR, "C_RESUMEN.md")

M = Matcher.desde_archivo()
REGISTROS = []
# contadores de diagnóstico
STATS = {}  # fuente -> dict(filas, match, por_metodo, sin_match_ejemplos)


# ── utilidades ──────────────────────────────────────────────────────────────
def _s(v):
    return "" if v is None else str(v).strip()


def es_cedula_candidata(v):
    """True si el string parece cédula colombiana (no un celular)."""
    c = norm_cedula(v)
    if not (6 <= len(c) <= 10):
        return False
    if len(c) == 10 and c.startswith("3"):  # celular
        return False
    return True


def si_no(v):
    s = _s(v).lower()
    if s in ("true", "si", "sí", "sí ", "x", "1", "1.0"):
        return "sí"
    if s in ("false", "no", "0", "0.0"):
        return "no"
    return None


def abrir(rel, sheet):
    from openpyxl import load_workbook
    p = os.path.join(RAIZ, rel)
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    filas = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return filas


def buscar_header(filas, tokens, limite=12):
    """Índice de la primera fila (en las primeras `limite`) que contenga algún token."""
    toks = [t.lower() for t in tokens]
    for i, fila in enumerate(filas[:limite]):
        celdas = " | ".join(_s(c).lower() for c in fila)
        if any(t in celdas for t in toks):
            return i
    return None


def idx_col(header, *patrones):
    """Índice de la primera columna cuyo nombre matchee algún patrón regex (i)."""
    for pat in patrones:
        rx = re.compile(pat, re.I)
        for j, c in enumerate(header):
            if rx.search(_s(c)):
                return j
    return None


def emitir(canon, metodo, campo, valor, fuente, hoja, anio, ced_fuente, nombre_crudo):
    REGISTROS.append({
        "canon": canon,
        "cedula_fuente": ced_fuente or "",
        "campo": campo,
        "valor": valor,
        "fuente_archivo": fuente,
        "hoja": hoja,
        "anio": anio,
        "metodo_match": metodo,
        "nombre_crudo": nombre_crudo or "",
    })


def registrar_stat(fuente, matched, metodo, crudo):
    st = STATS.setdefault(fuente, {"filas": 0, "match": 0,
                                   "por_metodo": {"cedula": 0, "email": 0, "nombre": 0},
                                   "sin_match": []})
    st["filas"] += 1
    if matched:
        st["match"] += 1
        st["por_metodo"][metodo] += 1
    elif len(st["sin_match"]) < 5:
        st["sin_match"].append(crudo)


def procesar_persona(fuente, hoja, anio, campos, cedula=None, email=None, nombre=None):
    """campos: lista de (campo, valor). Matchea y emite un registro por campo."""
    canon, metodo = M.match(cedula=cedula, email=email, nombre=nombre)
    crudo = _s(nombre) or _s(email) or _s(cedula)
    registrar_stat(fuente, canon is not None, metodo, crudo)
    if canon is None:
        return
    ced_f = norm_cedula(cedula) if cedula else ""
    for campo, valor in campos:
        if valor in (None, "", "nan"):
            continue
        emitir(canon, metodo, campo, valor, fuente, hoja, anio, ced_f, crudo)


# ── FAMILIA 1: 2021 presentación proyecto final (dos bloques por hoja) ────────
def familia_2021():
    archivos = [
        ("3-2021\\Presentación Proyecto Final G1MED.xlsx", "G1MED"),
        ("3-2021\\Presentación Proyecto Final G2MED.xlsx", "G2MED"),
        ("3-2021\\Presentación Proyecto Final G3ENV.xlsx", "G3ENV"),
    ]
    for rel, sheet in archivos:
        base = os.path.basename(rel)
        filas = abrir(rel, sheet)
        h = buscar_header(filas, ["Nombre"])
        if h is None:
            continue
        header = [_s(c) for c in filas[h]]
        # dos bloques: [Cita,Nombre,Apellido,GRUPO] cols aprox 0-3 y 5-8
        bloques = []
        # detectar columnas Nombre / Apellido / GRUPO en cada bloque
        n_idx = [j for j, c in enumerate(header) if c.lower().startswith("nombre")]
        for ni in n_idx:
            ap = next((j for j in range(ni, min(ni + 3, len(header)))
                       if header[j].lower().startswith("apellido")), None)
            gr = next((j for j in range(ni, min(ni + 4, len(header)))
                       if header[j].lower().startswith("grupo")), None)
            bloques.append((ni, ap, gr))
        for fila in filas[h + 1:]:
            for ni, ap, gr in bloques:
                nom = _s(fila[ni]) if ni < len(fila) else ""
                apell = _s(fila[ap]) if ap is not None and ap < len(fila) else ""
                grupo = _s(fila[gr]) if gr is not None and gr < len(fila) else ""
                nombre = (nom + " " + apell).strip()
                if not nombre or len(nombre) < 3:
                    continue
                campos = [("presento_proyecto", "sí")]
                if grupo:
                    campos.append(("mesa_categoria", grupo))
                procesar_persona(base, sheet, 2021, campos, nombre=nombre)


# ── FAMILIA 2: 2023 Presentación de Proyectos Finales (cronogramas variados) ──
def familia_2023_presentacion():
    rel = "5-2023\\Presentación de Proyectos Finales.xlsx"
    base = os.path.basename(rel)
    from openpyxl import load_workbook
    wb = load_workbook(os.path.join(RAIZ, rel), read_only=True)
    hojas = [s for s in wb.sheetnames if "parámetros" not in s.lower()
             and "parametros" not in s.lower()]
    wb.close()
    for sheet in hojas:
        filas = abrir(rel, sheet)
        h = buscar_header(filas, ["NOMBRE COMPLETO", "Participante"])
        if h is None:
            continue
        header = [_s(c) for c in filas[h]]
        col_nom = idx_col(header, r"nombre\s*completo", r"participante")
        if col_nom is None:
            continue
        col_ciudad = idx_col(header, r"^ciudad$")
        col_mentor = idx_col(header, r"mentor")
        # las filas de datos empiezan tras un posible sub-header (fila con AVANZADO/SI/NO)
        for fila in filas[h + 1:]:
            nombre = _s(fila[col_nom]) if col_nom < len(fila) else ""
            if not nombre or len(nombre) < 3:
                continue
            if nombre.lower() in ("nombre completo", "avanzado", "basico", "básico"):
                continue
            # buscar email y cédula en toda la fila
            email = ""
            cedula = ""
            for c in fila:
                s = _s(c)
                if not email and norm_email(s):
                    email = norm_email(s)
                if not cedula and es_cedula_candidata(s):
                    cedula = norm_cedula(s)
            ciudad = _s(fila[col_ciudad]) if col_ciudad is not None and col_ciudad < len(fila) else ""
            campos = [("presento_proyecto", "sí")]
            if ciudad:
                campos.append(("mesa_categoria", ciudad))
            procesar_persona(base, sheet, 2023, campos,
                             cedula=cedula or None, email=email or None, nombre=nombre)


# ── FAMILIA 3: 2023 Expoferia por mesa ────────────────────────────────────────
def familia_2023_expoferia():
    rel = "5-2023\\Jóvenes que presentan proyecto Expoferia por mesa.xlsx"
    base = os.path.basename(rel)
    sheet = "Presentación de Pagina web"
    filas = abrir(rel, sheet)
    h = buscar_header(filas, ["Participante"])
    if h is None:
        return
    header = [_s(c) for c in filas[h]]
    c_part = idx_col(header, r"participante")
    c_mesa = idx_col(header, r"^mesa$")
    c_grupo = idx_col(header, r"^grupo$")
    for fila in filas[h + 1:]:
        nombre = _s(fila[c_part]) if c_part < len(fila) else ""
        if not nombre or len(nombre) < 3:
            continue
        mesa = _s(fila[c_mesa]) if c_mesa is not None and c_mesa < len(fila) else ""
        grupo = _s(fila[c_grupo]) if c_grupo is not None and c_grupo < len(fila) else ""
        etiqueta = ("Mesa " + mesa.replace(".0", "") if mesa else "") + (" " + grupo if grupo else "")
        campos = [("presento_proyecto", "sí")]
        if etiqueta.strip():
            campos.append(("mesa_categoria", etiqueta.strip()))
        procesar_persona(base, sheet, 2023, campos, nombre=nombre)


# ── FAMILIA 4: 2024 NOTA AVANZADOS / NOTA BÁSICOS (nota_final = TOTAL) ─────────
def familia_2024_notas():
    rel = "6-2024\\Copia de Copia de Presentación de proyectos finales Jóvenes creaTIvos 2024.xlsx"
    base = os.path.basename(rel)
    for sheet in ("NOTA AVANZADOS", "NOTA BÁSICOS"):
        filas = abrir(rel, sheet)
        h = buscar_header(filas, ["Nombres", "E-mail", "TOTAL"])
        if h is None:
            continue
        header = [_s(c) for c in filas[h]]
        c_nom = idx_col(header, r"^nombres$")
        c_ape = idx_col(header, r"^apellidos$")
        c_email = idx_col(header, r"e-?mail")
        c_id = idx_col(header, r"^id$")
        c_total = idx_col(header, r"^total$")
        c_grupo = idx_col(header, r"^grupo$")
        c_obs = idx_col(header, r"observaci")
        for fila in filas[h + 1:]:
            nom = _s(fila[c_nom]) if c_nom is not None and c_nom < len(fila) else ""
            ape = _s(fila[c_ape]) if c_ape is not None and c_ape < len(fila) else ""
            nombre = (nom + " " + ape).strip()
            email = norm_email(fila[c_email]) if c_email is not None and c_email < len(fila) else ""
            cedula = norm_cedula(fila[c_id]) if c_id is not None and c_id < len(fila) else ""
            if not nombre and not email and not cedula:
                continue
            total = _s(fila[c_total]) if c_total is not None and c_total < len(fila) else ""
            grupo = _s(fila[c_grupo]) if c_grupo is not None and c_grupo < len(fila) else ""
            obs = _s(fila[c_obs]) if c_obs is not None and c_obs < len(fila) else ""
            try:
                total_num = float(total.replace(",", "."))
            except ValueError:
                total_num = None
            campos = []
            # presentó: TOTAL>0 o texto de observación que confirme; "no se presentó" -> no
            if obs and re.search(r"no se present", obs, re.I):
                campos.append(("presento_proyecto", "no"))
            elif total_num is not None and total_num > 0:
                campos.append(("presento_proyecto", "sí"))
            if total_num is not None:
                campos.append(("nota_final", total_num))
            if grupo:
                campos.append(("mesa_categoria", grupo))
            if obs:
                campos.append(("observacion", obs))
            if not campos:
                continue
            procesar_persona(base, sheet, 2024, campos,
                             cedula=cedula or None, email=email or None, nombre=nombre or None)


# ── FAMILIA 5: 2024 Distribución de horarios (¿Se conectó?) ───────────────────
def familia_2024_distribucion():
    rel = "6-2024\\Copia de Copia de Presentación de proyectos finales Jóvenes creaTIvos 2024.xlsx"
    base = os.path.basename(rel)
    sheet = "Distribución de horarios"
    filas = abrir(rel, sheet)
    h = buscar_header(filas, ["Nombre completo", "¿Se conectó?"])
    if h is None:
        return
    header = [_s(c) for c in filas[h]]
    c_nom = idx_col(header, r"nombre\s*completo")
    c_email = idx_col(header, r"correo\s*particip")
    c_id = idx_col(header, r"^id$")
    c_grupo = idx_col(header, r"^grupo$")
    c_con = idx_col(header, r"se\s*conect")
    for fila in filas[h + 1:]:
        nombre = _s(fila[c_nom]) if c_nom is not None and c_nom < len(fila) else ""
        email = norm_email(fila[c_email]) if c_email is not None and c_email < len(fila) else ""
        cedula = norm_cedula(fila[c_id]) if c_id is not None and c_id < len(fila) else ""
        if not nombre and not email and not cedula:
            continue
        grupo = _s(fila[c_grupo]) if c_grupo is not None and c_grupo < len(fila) else ""
        con = si_no(fila[c_con]) if c_con is not None and c_con < len(fila) else None
        campos = []
        if con:
            campos.append(("presento_proyecto", con))
        if grupo:
            campos.append(("mesa_categoria", grupo))
        if not campos:
            continue
        procesar_persona(base, sheet, 2024, campos,
                         cedula=cedula or None, email=email or None, nombre=nombre or None)


# ── FAMILIA 6: 2022 Encuesta Final BAQ (col "Presentación de Proyecto") ───────
def familia_2022_baq():
    rel = "4-2022\\Encuesta Final y Asistencia Clausura JC2022 BAQ.xlsx"
    base = os.path.basename(rel)
    filas = abrir(rel, None)
    h = buscar_header(filas, ["Nombre Completo", "Número de documento"])
    if h is None:
        return
    header = [_s(c) for c in filas[h]]
    c_nom = idx_col(header, r"nombre\s*completo")
    c_email = idx_col(header, r"correo")
    c_doc = idx_col(header, r"n[uú]mero de documento")
    c_grupo = idx_col(header, r"grupo")
    c_pp = idx_col(header, r"presentaci[oó]n de proyecto")
    for fila in filas[h + 1:]:
        nombre = _s(fila[c_nom]) if c_nom is not None and c_nom < len(fila) else ""
        email = norm_email(fila[c_email]) if c_email is not None and c_email < len(fila) else ""
        cedula = norm_cedula(fila[c_doc]) if c_doc is not None and c_doc < len(fila) else ""
        if not nombre and not email and not cedula:
            continue
        pp = si_no(fila[c_pp]) if c_pp is not None and c_pp < len(fila) else None
        grupo = _s(fila[c_grupo]) if c_grupo is not None and c_grupo < len(fila) else ""
        campos = []
        if pp:
            campos.append(("presento_proyecto", pp))
        if grupo:
            campos.append(("mesa_categoria", grupo))
        if not campos:
            continue
        procesar_persona(base, sheet_or("Respuestas de formulario 1"), 2022, campos,
                         cedula=cedula or None, email=email or None, nombre=nombre or None)


def sheet_or(x):
    return x


# ── FAMILIA 7: 2022 Encuesta Final CONSOLIDADA (col "PROYECTO FINAL") ─────────
def familia_2022_consolidada():
    rel = "4-2022\\Encuesta Final JC2022 CONSOLIDADA.xlsx"
    base = os.path.basename(rel)
    sheet = "Consolidado"
    filas = abrir(rel, sheet)
    h = buscar_header(filas, ["NOMBRE COMPLETO", "# DOCUMENTO"])
    if h is None:
        return
    header = [_s(c) for c in filas[h]]
    c_nom = idx_col(header, r"nombre\s*completo")
    c_email = idx_col(header, r"correo")
    c_doc = idx_col(header, r"#\s*documento")
    c_grupo = idx_col(header, r"^grupo$")
    c_pf = idx_col(header, r"^proyecto final$")
    for fila in filas[h + 1:]:
        nombre = _s(fila[c_nom]) if c_nom is not None and c_nom < len(fila) else ""
        email = norm_email(fila[c_email]) if c_email is not None and c_email < len(fila) else ""
        cedula = norm_cedula(fila[c_doc]) if c_doc is not None and c_doc < len(fila) else ""
        if not nombre and not email and not cedula:
            continue
        pf = si_no(fila[c_pf]) if c_pf is not None and c_pf < len(fila) else None
        grupo = _s(fila[c_grupo]) if c_grupo is not None and c_grupo < len(fila) else ""
        campos = []
        if pf:
            campos.append(("presento_proyecto", pf))
        if grupo:
            campos.append(("mesa_categoria", grupo))
        if not campos:
            continue
        procesar_persona(base, sheet, 2022, campos,
                         cedula=cedula or None, email=email or None, nombre=nombre or None)


# ── FAMILIA 8: 2019 PUNTUACION FINALISTAS (puntaje de SELECCIÓN, no proyecto) ──
def familia_2019():
    rel = "1-2019\\PUNTUACION FINALISTAS - Registro, puntos jurado y tecnico.xlsx"
    base = os.path.basename(rel)
    for sheet in ("CONFIRMACIÓN LLAMADAS", "Copia de CONFIRMACIÓN LLAMADAS"):
        filas = abrir(rel, sheet)
        h = buscar_header(filas, ["CEDULA", "PUNTAJE FASE 2"])
        if h is None:
            continue
        header = [_s(c) for c in filas[h]]
        c_nom = idx_col(header, r"^nombre$")
        c_ape = idx_col(header, r"^apellido$")
        c_email = idx_col(header, r"^email$")
        c_ced = idx_col(header, r"^cedula$")
        c_f2 = idx_col(header, r"puntaje fase 2")
        for fila in filas[h + 1:]:
            nom = _s(fila[c_nom]) if c_nom is not None and c_nom < len(fila) else ""
            ape = _s(fila[c_ape]) if c_ape is not None and c_ape < len(fila) else ""
            nombre = (nom + " " + ape).strip()
            email = norm_email(fila[c_email]) if c_email is not None and c_email < len(fila) else ""
            cedula = norm_cedula(fila[c_ced]) if c_ced is not None and c_ced < len(fila) else ""
            f2 = _s(fila[c_f2]) if c_f2 is not None and c_f2 < len(fila) else ""
            if not (nombre or email or cedula) or not f2:
                continue
            try:
                val = float(f2.replace(",", "."))
            except ValueError:
                continue
            procesar_persona(base, sheet, 2019,
                             [("puntaje_seleccion_fase2", val)],
                             cedula=cedula or None, email=email or None, nombre=nombre or None)


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(SALIDA_DIR, exist_ok=True)
    print("[C] extrayendo resultados / proyecto final / puntajes ...")
    familia_2021()
    familia_2023_presentacion()
    familia_2023_expoferia()
    familia_2024_notas()
    familia_2024_distribucion()
    familia_2022_baq()
    familia_2022_consolidada()
    familia_2019()

    with open(PAYLOAD, "w", encoding="utf-8") as fh:
        json.dump(REGISTROS, fh, ensure_ascii=False, indent=2)
    print(f"[C] payload → {PAYLOAD}  ({len(REGISTROS)} registros)")

    escribir_resumen()
    print(f"[C] resumen → {RESUMEN}")


def escribir_resumen():
    # agregados
    por_campo = {}
    por_anio_campo = {}
    canon_con_proyecto = set()
    canon_con_nota = set()
    metodo_global = {"cedula": 0, "email": 0, "nombre": 0}
    for r in REGISTROS:
        por_campo[r["campo"]] = por_campo.get(r["campo"], 0) + 1
        k = (r["anio"], r["campo"])
        por_anio_campo[k] = por_anio_campo.get(k, 0) + 1
        if r["campo"] == "presento_proyecto" and r["valor"] == "sí":
            canon_con_proyecto.add(r["canon"])
        if r["campo"] == "nota_final":
            canon_con_nota.add(r["canon"])
        metodo_global[r["metodo_match"]] = metodo_global.get(r["metodo_match"], 0) + 1

    total_match_registros = sum(metodo_global.values())
    lines = []
    lines.append("# Clúster C — RESULTADOS / PROYECTO FINAL / PUNTAJES")
    lines.append("")
    lines.append("Extracción de fin de programa de Jóvenes creaTIvos (2019-2024). "
                 "PII: solo en `tools/`. Payload: `C_resultados.json`.")
    lines.append("")
    lines.append(f"- **Registros totales (campo-nivel):** {len(REGISTROS)}")
    lines.append(f"- **Personas distintas con proyecto presentado (sí):** {len(canon_con_proyecto)}")
    lines.append(f"- **Personas distintas con nota_final (2024):** {len(canon_con_nota)}")
    lines.append("")
    lines.append("## Registros por campo")
    lines.append("")
    lines.append("| campo | registros |")
    lines.append("|---|---|")
    for campo, n in sorted(por_campo.items(), key=lambda x: -x[1]):
        lines.append(f"| {campo} | {n} |")
    lines.append("")
    lines.append("## Registros por año × campo")
    lines.append("")
    lines.append("| año | campo | registros |")
    lines.append("|---|---|---|")
    for (anio, campo), n in sorted(por_anio_campo.items()):
        lines.append(f"| {anio} | {campo} | {n} |")
    lines.append("")
    lines.append("## % de match por método (registros emitidos)")
    lines.append("")
    if total_match_registros:
        for met in ("cedula", "email", "nombre"):
            n = metodo_global.get(met, 0)
            pct = 100.0 * n / total_match_registros
            lines.append(f"- **{met}:** {n} ({pct:.1f}%)")
    lines.append("")
    lines.append("### Riesgo de homónimos (match por nombre)")
    lines.append("")
    n_nombre = metodo_global.get("nombre", 0)
    pct_nombre = 100.0 * n_nombre / total_match_registros if total_match_registros else 0
    lines.append(f"El **{pct_nombre:.1f}%** de los registros se ancló SOLO por nombre "
                 "(fuentes 2021 y buena parte de 2023 no traen cédula ni correo). "
                 "El match por nombre normaliza sin tildes y ordena tokens, por lo que "
                 "**homónimos reales pueden colisionar** (falsos positivos). Cada registro "
                 "conserva `nombre_crudo` para auditar. Recomendación: tratar `presento_proyecto` "
                 "resuelto por nombre como *probable*, no como verdad dura, hasta cruzar con "
                 "cédula/correo de otra fuente.")
    lines.append("")
    lines.append("## Cobertura y match por fuente")
    lines.append("")
    lines.append("| fuente | filas leídas | matcheadas | ced | email | nombre | ejemplos sin match |")
    lines.append("|---|---|---|---|---|---|---|")
    for fuente, st in sorted(STATS.items()):
        pm = st["por_metodo"]
        ej = "; ".join(st["sin_match"][:3]) if st["sin_match"] else "—"
        ej = ej.replace("|", "/")[:60]
        lines.append(f"| {fuente} | {st['filas']} | {st['match']} | "
                     f"{pm['cedula']} | {pm['email']} | {pm['nombre']} | {ej} |")
    lines.append("")
    lines.append("## Notas honestas por fuente (qué SÍ y qué NO trae)")
    lines.append("")
    lines.append("- **2019 · PUNTUACION FINALISTAS**: pese al nombre del archivo "
                 "(\"puntos jurado y tecnico\"), las columnas son de la **fase de SELECCIÓN** "
                 "(PROMEDIO, PUNTAJE PRUEBA, PREGUNTAS, PUNTAJE FASE 2), no del proyecto final. "
                 "Se emite como `puntaje_seleccion_fase2` (campo distinto) para no confundirlo "
                 "con resultados de fin de programa. Solo ~25 finalistas.")
    lines.append("- **2021 · Presentación Proyecto Final G1/G2/G3**: cronogramas de "
                 "presentación (Nombre+Apellido+GRUPO). Dan `presento_proyecto=sí` + grupo, "
                 "**sin cédula ni correo** → match solo por nombre.")
    lines.append("- **2023 · Presentación de Proyectos Finales**: 11 hojas (cronogramas por "
                 "ciudad/día). Estructura irregular; algunas traen correo, otras cédula, otras "
                 "solo nombre. Dan `presento_proyecto=sí` + ciudad como mesa/categoría.")
    lines.append("- **2023 · Expoferia por mesa**: `presento_proyecto=sí` + Mesa/Grupo. Nombre.")
    lines.append("- **2024 · NOTA AVANZADOS / NOTA BÁSICOS**: única fuente con **nota numérica** "
                 "real (columna TOTAL) + observaciones. `nota_final`=TOTAL, `presento_proyecto` "
                 "derivado (TOTAL>0, o 'no se presentó' en observación). Match por cédula (ID).")
    lines.append("- **2024 · Distribución de horarios**: `¿Se conectó?` → `presento_proyecto`. "
                 "Match por cédula/correo.")
    lines.append("- **2022 · Encuesta Final BAQ**: columna 'Presentación de Proyecto' (SI/NO) "
                 "+ cédula. `presento_proyecto`.")
    lines.append("- **2022 · Encuesta Final CONSOLIDADA**: columna 'PROYECTO FINAL' (SI/NO) "
                 "+ # documento. `presento_proyecto`.")
    lines.append("")
    lines.append("## Qué NO existe en ninguna fuente del clúster")
    lines.append("")
    lines.append("- **`nombre_proyecto`**: las listas son cronogramas, no catálogos de proyectos.")
    lines.append("- **`puntaje_jurado` / `puntaje_tecnico` separados**: no hay rúbricas con "
                 "columnas jurado/técnico. Lo más cercano es `nota_final` (TOTAL 2024) y el "
                 "`puntaje_seleccion_fase2` de 2019 (que es de selección, no de proyecto).")
    lines.append("")

    with open(RESUMEN, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


if __name__ == "__main__":
    main()
