# -*- coding: utf-8 -*-
"""
enriquecimiento_helper.py — Utilidades compartidas para los subagentes de la etapa
final de enriquecimiento (clústeres A–F). Evita que cada agente reimplemente lectura
de xlsx y, sobre todo, el "mejor match" de una fila histórica → una persona conocida.

Provee:
  1. Normalizadores robustos: norm_cedula / norm_email / norm_nombre (sin tildes, lower).
  2. Roster de identidad: exporta de Supabase (participants + postulantes_jc/mr) los
     índices por cédula, correo y nombre → `tools/enriquecimiento/roster_objetivo.json`.
  3. Matcher: dada (cedula?, email?, nombre?) devuelve (cedula_canonica, metodo) con
     prioridad CÉDULA > CORREO > NOMBRE. Así una fuente que solo trae correo o nombre
     igual se ancla a la persona correcta.
  4. Lectura de hojas Excel/CSV con detección de fila de encabezado (streaming).

Uso como script:
    python enriquecimiento_helper.py --exportar-roster   # regenera el roster desde Supabase
Uso como módulo (en los scripts de cada clúster):
    from enriquecimiento_helper import Matcher, cargar_hoja, norm_cedula, norm_email
    m = Matcher.desde_archivo()          # carga tools/enriquecimiento/roster_objetivo.json
    ced, metodo = m.match(cedula=fila.get("ID"), email=fila.get("E-mail"), nombre=...)

PII: el roster y los payloads viven SOLO en tools/ (gitignoreado). Nunca a docs/ ni al repo.
"""

import argparse
import json
import os
import re
import sys
import unicodedata

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
PROYECTO_ROOT = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
ROSTER_PATH = os.path.join(PROYECTO_ROOT, "tools", "enriquecimiento", "roster_objetivo.json")

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = __import__("io").TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


# ── Normalizadores ────────────────────────────────────────────────────────────
def norm_cedula(v):
    """Cédula → dígitos puros (quita '.0', espacios, puntos de miles, guiones)."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"[^\d]", "", s)
    return s.lstrip("0") or s  # quita ceros a la izquierda pero no deja cadena vacía si era '0...'


def norm_email(v):
    if v is None:
        return ""
    s = str(v).strip().lower()
    m = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", s)
    return m.group(0) if m else ""


def _sin_tildes(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def norm_nombre(v):
    """Nombre → sin tildes, minúsculas, tokens ordenados (robusto a orden nombre/apellido)."""
    if v is None:
        return ""
    s = _sin_tildes(str(v)).lower()
    s = re.sub(r"[^a-z\s]", " ", s)
    tokens = [t for t in s.split() if len(t) > 1]
    return " ".join(sorted(tokens))


# ── Roster de identidad ─────────────────────────────────────────────────────────
def _pick(row, *patrones):
    """Devuelve el primer valor de row cuya clave matchee alguno de los patrones (regex, i)."""
    for pat in patrones:
        rx = re.compile(pat, re.I)
        for k, val in row.items():
            if rx.search(k):
                return val
    return None


def exportar_roster(salida=ROSTER_PATH):
    """Construye índices cedula/email/nombre desde participants + postulantes_jc/mr."""
    sys.path.insert(0, DIRECTORIO_SCRIPT)
    from cargar_supabase import Supa, cargar_env_local
    cargar_env_local()
    supa = Supa(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_ROLE_KEY"])

    by_cedula, by_email, by_nombre = {}, {}, {}
    personas = []

    def _ingresar(cedula, email, nombre, origen):
        c = norm_cedula(cedula)
        e = norm_email(email)
        n = norm_nombre(nombre)
        if not c and not e:
            return
        canon = c or ("email:" + e)  # clave canónica: cédula si hay, si no el correo
        personas.append({"canon": canon, "cedula": c, "email": e, "nombre": n, "origen": origen})
        if c:
            by_cedula.setdefault(c, canon)
        if e:
            by_email.setdefault(e, canon)
        if n:
            by_nombre.setdefault(n, canon)

    for row in supa.get_todo("/participants?select=q10_id,email,nombre"):
        _ingresar(row.get("q10_id"), row.get("email"), row.get("nombre"), "participants")
    for tabla in ("postulantes_jc", "postulantes_mr"):
        try:
            for row in supa.get_todo(f"/{tabla}?select=*"):
                _ingresar(_pick(row, r"cedula|documento|identific"),
                          _pick(row, r"email|correo"),
                          _pick(row, r"nombre"),
                          tabla)
        except Exception as e:
            print(f"[helper] aviso: no se pudo leer {tabla}: {e}")

    os.makedirs(os.path.dirname(salida), exist_ok=True)
    with open(salida, "w", encoding="utf-8") as fh:
        json.dump({"by_cedula": by_cedula, "by_email": by_email, "by_nombre": by_nombre,
                   "n_personas": len(personas)}, fh, ensure_ascii=False)
    print(f"[helper] roster → {salida}  (cédulas={len(by_cedula)} correos={len(by_email)} "
          f"nombres={len(by_nombre)} filas={len(personas)})")


class Matcher:
    def __init__(self, by_cedula, by_email, by_nombre):
        self.by_cedula, self.by_email, self.by_nombre = by_cedula, by_email, by_nombre

    @classmethod
    def desde_archivo(cls, path=ROSTER_PATH):
        with open(path, encoding="utf-8") as fh:
            d = json.load(fh)
        return cls(d["by_cedula"], d["by_email"], d["by_nombre"])

    def match(self, cedula=None, email=None, nombre=None):
        """Devuelve (canon, metodo) con prioridad cédula > correo > nombre; (None, None) si no."""
        c = norm_cedula(cedula)
        if c and c in self.by_cedula:
            return self.by_cedula[c], "cedula"
        e = norm_email(email)
        if e and e in self.by_email:
            return self.by_email[e], "email"
        n = norm_nombre(nombre)
        if n and n in self.by_nombre:
            return self.by_nombre[n], "nombre"
        return None, None


# ── Lectura de hojas (header auto-detectado) ────────────────────────────────────
def _norm_celda(v):
    from datetime import date, datetime
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip()


def _detectar_header(filas):
    mejor_i, mejor = 0, -1
    for i, fila in enumerate(filas):
        celdas = [_norm_celda(c) for c in fila]
        cortas = [c for c in celdas if c and len(c) <= 60 and not re.fullmatch(r"-?\d+([.,]\d+)?", c)]
        score = len(cortas) + 0.3 * len([c for c in celdas if c])
        if score > mejor:
            mejor_i, mejor = i, score
    return mejor_i


def cargar_hoja(path, sheet=None, escaneo_header=15, limite=None):
    """Devuelve (columnas, filas_dict). filas_dict: lista de dicts col→valor.
    Detecta la fila de encabezado en las primeras `escaneo_header` filas."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        filas = _consumir_iter(it, escaneo_header, limite)
        wb.close()
        return filas
    else:  # csv
        import csv
        for enc in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(path, encoding=enc, newline="") as fh:
                    try:
                        sep = csv.Sniffer().sniff(fh.read(4096), delimiters=",;\t|").delimiter
                    except Exception:
                        sep = ","
                    fh.seek(0)
                    return _consumir_iter(csv.reader(fh, delimiter=sep), escaneo_header, limite)
            except UnicodeDecodeError:
                continue
    return [], []


def _consumir_iter(it, escaneo_header, limite):
    buffer = []
    for fila in it:
        buffer.append(fila)
        if len(buffer) >= escaneo_header:
            break
    if not buffer:
        return [], []
    h = _detectar_header(buffer)
    columnas = [_norm_celda(c) for c in buffer[h]]
    # desambiguar
    vistos = {}
    cols = []
    for j, c in enumerate(columnas):
        base = c or f"col_{j+1}"
        vistos[base] = vistos.get(base, 0) + 1
        cols.append(base if vistos[base] == 1 else f"{base}__{vistos[base]}")
    filas = []
    for fila in buffer[h + 1:]:
        filas.append(fila)
    def _emit(fila):
        celdas = list(fila)
        return {cols[j]: celdas[j] if j < len(celdas) else None for j in range(len(cols))}
    salida = []
    for fila in filas:
        if any(_norm_celda(c) for c in list(fila)[:len(cols)]):
            salida.append(_emit(fila))
        if limite and len(salida) >= limite:
            return cols, salida
    for fila in it:
        if any(_norm_celda(c) for c in list(fila)[:len(cols)]):
            salida.append(_emit(fila))
        if limite and len(salida) >= limite:
            break
    return cols, salida


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--exportar-roster", action="store_true")
    args = ap.parse_args()
    if args.exportar_roster:
        exportar_roster()
    else:
        ap.print_help()
