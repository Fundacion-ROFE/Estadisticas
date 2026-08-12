# -*- coding: utf-8 -*-
"""
extraer_avance_monitorias_2024.py
=================================
Clúster D — enriquecimiento JC 2024.

Hipótesis: Q10 purgó el % de avance de los retirados. La "BD de Seguimiento de
Monitorías JC2024" (fuente MANUAL, ajena a Q10) puede conservar ese avance para
los 139 retirados de la cohorte JC 2024 que no tienen ninguna fila de avance en
Supabase.

Qué hace:
  1. Recorre TODAS las hojas de ambos .xlsx de Seguimiento de Monitorías.
  2. Ubica de forma robusta la columna de cédula y las columnas de avance por
     curso (HB / EMP / LÓG / JS / HTML / CSS), tanto en hojas "anchas" con
     encabezado como en hojas por-curso con encabezados corridos (la 1ª fila de
     datos quedó como header).
  3. Normaliza cédula (quita ".0", espacios, puntos) y avance (0–1 -> 0–100;
     ignora vacíos / #N/A / textos como "avanzó").
  4. Cruza contra las 139 cédulas objetivo.

Salidas (PII -> tools/, gitignoreado; NUNCA a docs/ ni a Supabase):
  - tools/enriquecimiento/D_avance_monitorias_2024.json        (solo las 139)
  - tools/enriquecimiento/D_avance_monitorias_2024_todos.json  (todas las cédulas)
  - tools/enriquecimiento/RESUMEN.md

Es SOLO LECTURA sobre los .xlsx. No hay red. Reutilizable / idempotente.
"""
from __future__ import annotations
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

# --------------------------------------------------------------------------- #
# Configuración
# --------------------------------------------------------------------------- #
FUENTES = [
    # Preferir la ACTUALIZADA (más reciente); la otra complementa/deduplica.
    r"C:\Users\EstudiantesJC\Downloads\COMPLETE-ORDEN-INFORMATION\6-2024\Copia de BD Seguimiento de Monitorias JC2024 - ACTUALIZADA_.xlsx",
    r"C:\Users\EstudiantesJC\Downloads\COMPLETE-ORDEN-INFORMATION\6-2024\Copia de BD Seguimiento de Monitorias JC2024 .xlsx",
]

OUT_DIR = Path(r"C:\Users\EstudiantesJC\downloads\admin-usable\tools\enriquecimiento")

CURSOS = ("HB", "EMP", "LOG", "JS", "HTML", "CSS")  # LÓG -> LOG (canónico)

CEDULAS_139 = {
    "1000443482","1000694698","1000778233","1001173352","1002192823","1002390952",
    "1007978497","1013340878","1013343323","1013460946","1013585551","1016950375",
    "1020105538","1021807486","1022003269","1022149634","1025529813","1025647178",
    "1025888455","1026554243","1027522661","1028780295","1028860796","1028861118",
    "1031650046","1031654847","1032013456","1032182072","1032370623","1033183375",
    "1039447343","1042250936","1042849202","1042969312","1043295998","1043299454",
    "1043299790","1043300665","1043302805","1043306006","1043643056","1043658584",
    "1043963959","1044611950","1044614339","1044620721","1044621906","1047375626",
    "1048442110","1048444398","1050277522","1052956148","1053123478","1061727276",
    "1063147644","1065589329","1066351765","1067599401","1069643219","1069715983",
    "1075598506","1092534626","1104413677","1104807134","1104936366","1105057716",
    "1105371993","1107060918","1107843105","1108254352","1108335924","1108561877",
    "1108562383","1108564425","1108644553","1109187674","1109188462","1109191058",
    "1109663200","1109667893","1109669419","1110040689","1111660714","1111669915",
    "1112049432","1114543521","1126594128","1127580354","1128045572","1129844360",
    "1130144635","1139427411","1140916180","1142921697","1143343234","1149934857",
    "1282273","1941796","48657419","49783017","50422080","5097622","51677002",
    "52617237","5336645","53369166","53451420","53771911","53816987","54190859",
    "54536665","54561539","55174816","55257397","55785025","55903986","56191811",
    "56344129","56497112","56522282","56603709","56629931","56807812","57048277",
    "57161249","57199234","5934628","6038199","7728432","930264940","930343793",
    "931548820","944081140","951813369","952163491","953601010","954900163",
    "955127345","966552259",
}

# --------------------------------------------------------------------------- #
# Helpers de normalización
# --------------------------------------------------------------------------- #
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def norm_txt(v) -> str:
    if v is None:
        return ""
    return _strip_accents(str(v)).strip().upper()


def curso_de_header(h) -> str | None:
    """Mapea un encabezado de columna a un curso canónico, o None."""
    n = norm_txt(h)
    if not n:
        return None
    base = re.sub(r"[\s\.]+", "", n)          # 'CEL + PAIS' -> 'CEL+PAIS'
    base = re.sub(r"\d+$", "", base)          # 'JS1' -> 'JS', 'HTML7' -> 'HTML'
    return base if base in CURSOS else None


def es_hoja_leccion(nombre: str) -> bool:
    """True si es una hoja de LECCIÓN/semana suelta (checkpoint), no de curso.

    'JS1','HTML7','CSS3','CCS2' -> lección (progreso de una sola lección, 0/1).
    'JS 8','HTML 8','CSS 8' (con espacio) -> hoja de curso -> NO es lección.
    Estas hojas de lección NO son avance a nivel de curso, así que se excluyen.
    """
    n = norm_txt(nombre)                      # conserva espacios internos
    return bool(re.fullmatch(r"(JS|HTML|CSS|CCS)\d+", n))


def curso_de_hoja(nombre: str) -> str | None:
    """Mapea el NOMBRE de una hoja por-curso (encabezado corrido) a un curso.

    Devuelve None para hojas de lección suelta (ver es_hoja_leccion)."""
    if es_hoja_leccion(nombre):
        return None
    n = re.sub(r"[\s\.]+", "", norm_txt(nombre))
    n = re.sub(r"\d+$", "", n)                # 'HB 8' -> 'HB', 'LÓG 8' -> 'LOG'
    return n if n in CURSOS else None


def es_col_id(h) -> bool:
    n = norm_txt(h)
    return n == "ID" or "DOCUMENTO" in n or "CEDULA" in n


def norm_cedula(v) -> str | None:
    """Devuelve la cédula como string de solo dígitos, o None si no aplica."""
    if v is None:
        return None
    if isinstance(v, float):
        if v != v:  # NaN
            return None
        v = int(v) if v.is_integer() else v
    s = str(v).strip()
    s = re.sub(r"[.\s,]", "", s)              # quita '.0', espacios, puntos, comas
    s = re.sub(r"\.0$", "", s)
    if not s.isdigit():
        return None
    if len(s) < 6 or len(s) > 10:            # cédulas/TI colombianas y ext.
        return None
    return s


def parse_avance(v):
    """0–1 -> 0–100 (float 1 decimal). None si no es un avance numérico válido."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if not re.fullmatch(r"-?\d+(\.\d+)?", s):
            return None
        v = float(s)
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f != f:            # NaN
        return None
    if f < 0 or f > 1.0001:   # avance en estas hojas viene 0–1
        return None
    return round(min(f, 1.0) * 100, 1)


# --------------------------------------------------------------------------- #
# Extracción
# --------------------------------------------------------------------------- #
def cargar_hojas(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        yield name, rows
    wb.close()


def detectar_header(rows):
    """Devuelve (idx_header, id_col, {curso: [cols]}) o (None, None, None)."""
    best = None
    for i in range(min(6, len(rows))):
        row = rows[i]
        cursos = defaultdict(list)
        id_col = None
        for j, cell in enumerate(row):
            c = curso_de_header(cell)
            if c:
                cursos[c].append(j)
            if id_col is None and es_col_id(cell):
                id_col = j
        score = len(cursos) + (1 if id_col is not None else 0)
        if best is None or score > best[0]:
            best = (score, i, id_col, dict(cursos))
    if best and best[0] >= 1:
        return best[1], best[2], best[3]
    return None, None, None


def extraer_ancha(rows, idx_header, id_col, cursos_cols, hoja, archivo):
    """Hoja con encabezado: id_col + columnas de curso (agrupa numeradas, max)."""
    registros = []
    for row in rows[idx_header + 1:]:
        if id_col >= len(row):
            continue
        ced = norm_cedula(row[id_col])
        if not ced:
            continue
        for curso, cols in cursos_cols.items():
            vals = [parse_avance(row[c]) for c in cols if c < len(row)]
            vals = [v for v in vals if v is not None]
            if not vals:
                continue
            registros.append({
                "cedula": ced, "curso": curso, "avance": max(vals),
                "hoja": hoja, "archivo": archivo, "fuente_tipo": "ancha",
            })
    return registros


def extraer_por_curso(rows, curso, master_cedulas, hoja, archivo):
    """Hoja por-curso con encabezado corrido: detecta col cédula y col avance."""
    ncols = max((len(r) for r in rows), default=0)
    if ncols == 0:
        return []
    # Columna cédula = candidata con mayor solape con el set maestro de cédulas.
    mejor_col, mejor_solape, mejor_cnt = None, -1, 0
    for j in range(ncols):
        vals = [norm_cedula(r[j]) for r in rows if j < len(r)]
        vals = [v for v in vals if v]
        if not vals:
            continue
        solape = sum(1 for v in vals if v in master_cedulas)
        if solape > mejor_solape or (solape == mejor_solape and len(vals) > mejor_cnt):
            mejor_col, mejor_solape, mejor_cnt = j, solape, len(vals)
    if mejor_col is None or mejor_solape <= 0:
        return []
    id_col = mejor_col
    # Columna avance = la de mayor densidad de valores en [0,1] (excluye id_col).
    av_col, av_dens = None, 0
    for j in range(ncols):
        if j == id_col:
            continue
        dens = sum(1 for r in rows if j < len(r) and parse_avance(r[j]) is not None)
        if dens > av_dens:
            av_col, av_dens = j, dens
    if av_col is None or av_dens == 0:
        return []
    registros = []
    for r in rows:
        if id_col >= len(r) or av_col >= len(r):
            continue
        ced = norm_cedula(r[id_col])
        av = parse_avance(r[av_col])
        if ced and av is not None:
            registros.append({
                "cedula": ced, "curso": curso, "avance": av,
                "hoja": hoja, "archivo": archivo, "fuente_tipo": "por_curso",
            })
    return registros


def main():
    # ---- Pass 1: set maestro de cédulas (de hojas con columna ID clara) ----
    master_cedulas = set()
    hojas_cache = []  # [(archivo, hoja, rows, idx_header, id_col, cursos_cols)]
    for path in FUENTES:
        archivo = Path(path).name
        for hoja, rows in cargar_hojas(path):
            if not rows:
                continue
            idx, id_col, cursos_cols = detectar_header(rows)
            if id_col is not None:
                for row in rows[(idx or 0) + 1:]:
                    if id_col < len(row):
                        c = norm_cedula(row[id_col])
                        if c:
                            master_cedulas.add(c)
            hojas_cache.append((archivo, hoja, rows, idx, id_col, cursos_cols))

    # ---- Pass 2: extracción de avance ----
    registros = []
    diag = []  # (archivo, hoja, tipo, n_registros)
    for archivo, hoja, rows, idx, id_col, cursos_cols in hojas_cache:
        recs = []
        if id_col is not None and cursos_cols:
            recs = extraer_ancha(rows, idx, id_col, cursos_cols, hoja, archivo)
            tipo = "ancha"
        else:
            curso = curso_de_hoja(hoja)
            if curso:
                recs = extraer_por_curso(rows, curso, master_cedulas, hoja, archivo)
                tipo = "por_curso"
            else:
                tipo = "skip"
        if recs:
            registros.extend(recs)
            diag.append((archivo, hoja, tipo, len(recs)))

    # ---- Dedupe por (cedula, curso): prioriza por_curso, luego mayor avance --
    def clave_prioridad(rec):
        # menor tupla = mejor: por_curso antes que ancha, luego avance desc
        return (0 if rec["fuente_tipo"] == "por_curso" else 1, -rec["avance"])

    mejor = {}
    for rec in registros:
        k = (rec["cedula"], rec["curso"])
        if k not in mejor or clave_prioridad(rec) < clave_prioridad(mejor[k]):
            mejor[k] = rec

    todos = sorted(mejor.values(), key=lambda r: (r["cedula"], r["curso"]))
    objetivo = [r for r in todos if r["cedula"] in CEDULAS_139]

    # ---- Escritura ----
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "D_avance_monitorias_2024.json").write_text(
        json.dumps(objetivo, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUT_DIR / "D_avance_monitorias_2024_todos.json").write_text(
        json.dumps(todos, ensure_ascii=False, indent=2), encoding="utf-8")

    # ---- Métricas ----
    ced_obj = {r["cedula"] for r in objetivo}
    por_curso_cnt = defaultdict(int)
    for r in objetivo:
        por_curso_cnt[r["curso"]] += 1
    en_master = len(CEDULAS_139 & master_cedulas)

    escribir_resumen(objetivo, ced_obj, por_curso_cnt, en_master, diag)

    # ---- Reporte a consola ----
    print("=== EXTRACCIÓN AVANCE MONITORÍAS 2024 ===")
    print(f"Cédulas maestro (todas las hojas): {len(master_cedulas)}")
    print(f"De las 139 objetivo, presentes en el archivo (cualquier hoja con ID): {en_master}")
    print(f"De las 139, RECUPERADAS con >=1 avance: {len(ced_obj)}")
    print(f"Registros (cedula,curso) para las 139: {len(objetivo)}")
    print("Desglose por curso (nº de cédulas objetivo con avance):")
    for c in CURSOS:
        print(f"   {c:5} {por_curso_cnt.get(c,0)}")
    print(f"\nSalidas en: {OUT_DIR}")


def escribir_resumen(objetivo, ced_obj, por_curso_cnt, en_master, diag):
    faltan = sorted(CEDULAS_139 - ced_obj)
    lines = []
    lines.append("# Clúster D — Avance de retirados JC 2024 desde BD Seguimiento de Monitorías\n")
    lines.append("Fuente MANUAL de monitores (ajena a Q10). Objetivo: recuperar el "
                 "% de avance que Q10 purgó para 139 retirados.\n")
    lines.append("## Resultado\n")
    lines.append(f"- Cédulas objetivo: **139**")
    lines.append(f"- Presentes en el archivo (alguna hoja con columna ID): **{en_master}**")
    lines.append(f"- **Recuperadas con >=1 valor de avance: {len(ced_obj)}**")
    lines.append(f"- Registros (cédula, curso) extraídos para las 139: **{len(objetivo)}**\n")
    lines.append("## Distribución por curso (nº de cédulas objetivo con avance)\n")
    lines.append("| Curso | Cédulas con avance |")
    lines.append("|---|---|")
    for c in CURSOS:
        lines.append(f"| {c} | {por_curso_cnt.get(c,0)} |")
    lines.append("")
    lines.append("## Ejemplos (primeras 15 filas del payload de las 139)\n")
    lines.append("| Cédula | Curso | Avance | Hoja | Archivo |")
    lines.append("|---|---|---|---|---|")
    for r in objetivo[:15]:
        lines.append(f"| {r['cedula']} | {r['curso']} | {r['avance']} | {r['hoja']} | {r['archivo']} |")
    lines.append("")
    lines.append(f"## Cédulas objetivo SIN avance recuperado ({len(faltan)})\n")
    lines.append(", ".join(faltan) if faltan else "_(ninguna)_")
    lines.append("")
    lines.append("## Hojas que aportaron registros (diagnóstico)\n")
    lines.append("| Archivo | Hoja | Tipo | Registros |")
    lines.append("|---|---|---|---|")
    for archivo, hoja, tipo, n in sorted(diag, key=lambda x: -x[3]):
        lines.append(f"| {archivo} | {hoja} | {tipo} | {n} |")
    lines.append("")
    (OUT_DIR / "RESUMEN.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
