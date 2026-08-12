# -*- coding: utf-8 -*-
"""
extraer_cluster_A_socioeconomico.py — CLÚSTER A (socioeconómico ampliado + convocatoria)

Recorre las fuentes históricas de Jóvenes creaTIvos (formularios de aplicación 2019-2025)
y extrae los campos SOCIOECONÓMICOS y de CONVOCATORIA que hoy NO existen en Supabase,
anclando cada fila a la persona conocida vía el Matcher del helper compartido
(prioridad cédula > correo > nombre).

Salidas (SOLO en tools/, gitignoreado, contiene PII):
  - tools/enriquecimiento/A_socioeconomico.json  (lista de registros)
  - tools/enriquecimiento/A_RESUMEN.md            (cobertura + ejemplos)

Cada registro: {canon, cedula_fuente, campo, valor, fuente_archivo, hoja, anio, metodo_match}
Un registro por (persona, campo, valor, fuente_archivo). Deduplica valores idénticos.

Idempotente: reejecutar reconstruye ambos archivos desde cero. No escribe a Supabase,
no modifica los .xlsx, no usa red ni truststore. Lectura openpyxl read_only vía el helper.
"""

import json
import os
import re
import sys
import unicodedata
from collections import defaultdict

DIRECTORIO_SCRIPT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, DIRECTORIO_SCRIPT)
from enriquecimiento_helper import Matcher, cargar_hoja, norm_cedula, norm_email, norm_nombre

PROYECTO_ROOT = os.path.abspath(os.path.join(DIRECTORIO_SCRIPT, "..", ".."))
RAIZ_FUENTES = r"C:\Users\EstudiantesJC\Downloads\COMPLETE-ORDEN-INFORMATION"
SALIDA_DIR = os.path.join(PROYECTO_ROOT, "tools", "enriquecimiento")
SALIDA_JSON = os.path.join(SALIDA_DIR, "A_socioeconomico.json")
SALIDA_RESUMEN = os.path.join(SALIDA_DIR, "A_RESUMEN.md")

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = __import__("io").TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


# ── Normalización de encabezados ────────────────────────────────────────────────
def _h(s):
    """Encabezado → sin tildes, minúsculas, espacios/saltos colapsados."""
    if s is None:
        return ""
    s = "".join(c for c in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(c))
    s = s.lower().replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()


# ── Patrones de campo objetivo ───────────────────────────────────────────────────
# (include_regex, exclude_regex).  Se escanean los encabezados normalizados por _h().
FIELD_PATTERNS = {
    "institucion_educativa": (r"instituci[o]n", r"tipo|adjunta"),
    "grado_escolar":        (r"\bgrado\b", r"promedio"),
    "promedio_colegio":     (r"promedio", r"ingreso|familiar|salario"),
    "nivel_ingles":         (r"ingl[e]s", None),
    "ingreso_hogar":        (r"ingreso", None),
    "personas_nucleo":      (r"(personas.*(nucleo|circulo|familiar))|(nucleo familiar)|(personas_nucleo)|(personas del nucleo)", None),
    "acceso_internet":      (r"internet", None),
    "acceso_computador":    (r"computador", None),
    "direccion":            (r"direcci[o]n", r"correo|e-?mail"),
    "barrio":               (r"barrio", None),
    "tipo_vivienda":        (r"(tipo de (casa|vivienda|hogar))|tipo_vivienda|(tipo de hogar)|(vivienda en la que habitas)|(f tipo de hogar)", None),
}

# Columnas de identidad para el match (no se emiten como campo).
CED_INC = r"(numero de identificaci|identificaci[o]n|documento|c[e]dula|^id$|ci ?/ ?di|^cc$|documento_norm|^ci ?/ ?di$)"
CED_EXC = r"tipo|adjunta|foto|imagen"
EMAIL_INC = r"correo|e-?mail"
NOMBRE_INC = r"nombre"
NOMBRE_EXC = r"instituci|usuario|final|concatenar|filtro"
APELLIDO_INC = r"apellido"
ANIO_INC = r"^(anio|a[nñ]o)$|^anio de|^a[nñ]o de"

VALORES_NULOS = {"", "n/a", "#n/a", "#ref!", "#value!", "-", ".", "na", "sin datos", "none", "null"}
NUMERICOS = {"grado_escolar", "promedio_colegio", "personas_nucleo"}


# Validadores de valor por campo. Devuelven True si el valor es plausible para el campo.
# Sirven de red de seguridad ante filas de origen desalineadas (p.ej. BD Inscritos
# BARRANQUILLA mezcla dos versiones del formulario con columnas corridas a partir de
# cierto punto). Solo se validan campos con vocabulario acotado o numérico.
def _v_acceso(h):
    return bool(re.match(r"^(si|no)\b", h)) or bool(re.search(
        r"internet|computador|cafe|prestad|personal|telefono|celular|adapto|tengo|linea|datos|movil", h))

def _v_ingles(h):
    return bool(re.search(r"basic|medio|avanzad|principiant|intermedi|no hablo|ninguno|alto|bajo|nativo|fluido|ingl", h))

def _v_ingreso(h):
    return bool(re.search(r"salario|minimo|sueldo|smmlv|ingreso", h))

def _v_personas(h):
    return bool(re.fullmatch(r"\d{1,2}", h)) or "persona" in h

def _v_grado(h):
    if re.search(r"bachiller|grado|ya soy|egresad|once|decim|primero|segundo|tercero|cuarto|quinto|sexto|septimo|octavo|noveno|universi|semestre|tecnic", h):
        return True
    m = re.match(r"(\d{1,2})", h)
    return bool(m) and 1 <= int(m.group(1)) <= 13

def _v_promedio(h):
    m = re.fullmatch(r"(\d{1,3})([.,]\d+)?", h)
    return bool(m) and 0 <= float(m.group(1)) <= 100

def _v_vivienda(h):
    return bool(re.search(r"propia|arrendad|familiar|compartid|prestad|hipotecad|invasion|usufructo", h))

VALIDATORS = {
    "acceso_internet": _v_acceso,
    "acceso_computador": _v_acceso,
    "nivel_ingles": _v_ingles,
    "ingreso_hogar": _v_ingreso,
    "personas_nucleo": _v_personas,
    "grado_escolar": _v_grado,
    "promedio_colegio": _v_promedio,
    "tipo_vivienda": _v_vivienda,
}


def _norm_val(field, v):
    if v is None:
        return None
    s = str(v).strip()
    if _h(s) in VALORES_NULOS:
        return None
    if field in NUMERICOS and re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    s = re.sub(r"\s+", " ", s)
    if not s:
        return None
    val = VALIDATORS.get(field)
    if val and not val(_h(s)):
        return None
    return s


def _primera_col(cols, include, exclude=None):
    rx_i = re.compile(include, re.I)
    rx_e = re.compile(exclude, re.I) if exclude else None
    for c in cols:
        hc = _h(c)
        if rx_i.search(hc) and not (rx_e and rx_e.search(hc)):
            return c
    return None


def construir_colmap(cols):
    """Devuelve (campos_map, id_cols). campos_map: campo->col. id_cols: dict de columnas identidad."""
    campos = {}
    for campo, (inc, exc) in FIELD_PATTERNS.items():
        col = _primera_col(cols, inc, exc)
        if col:
            campos[campo] = col
    ids = {
        "cedula": _primera_col(cols, CED_INC, CED_EXC),
        "email": _primera_col(cols, EMAIL_INC),
        "nombre": _primera_col(cols, NOMBRE_INC, NOMBRE_EXC),
        "apellido": _primera_col(cols, APELLIDO_INC),
        "anio": _primera_col(cols, ANIO_INC),
    }
    return campos, ids


def _sheetnames(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    nombres = list(wb.sheetnames)
    wb.close()
    return nombres


# ── Configuración de fuentes ─────────────────────────────────────────────────────
# dict: relpath, anio (None=por fila/desconocido), sheets (lista o None=AUTO),
#       estado_default (str|None), estado_col (regex de columna de convocatoria|None)
SOURCES = [
    # Consolidado (identidad + estado universal + socioeconómico disperso ~400)
    {"rel": "Jóvenes Creativos única tabla.xlsx", "anio": None, "sheets": ["Hoja1"],
     "estado_default": None, "estado_col": r"^estado$"},
    {"rel": "Jóvenes Creativos única tabla.xlsx", "anio": None, "sheets": ["Hoja2"],
     "estado_default": None, "estado_col": None},

    # 2025 — formularios Fase 1 (hoja limpia con encabezados nombrados)
    {"rel": r"7-2025\Copia de COL - Fase #1 Jóvenes creaTIvos 2025 (respuestas).xlsx", "anio": 2025,
     "sheets": ["Base con criterios de descarte"], "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"7-2025\Copia de ECU - Fase #1 Jóvenes creaTIvos 2025 (respuestas).xlsx", "anio": 2025,
     "sheets": ["Base con criterios de descarte"], "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"7-2025\Copia de Fase #1 Jóvenes creaTIvos 2025 UY (respuestas).xlsx", "anio": 2025,
     "sheets": ["Base con criterios de descarte"], "estado_default": "POSTULADO", "estado_col": None},

    # 2024
    {"rel": r"6-2024\Copia de BD JC 2024 General.xlsx", "anio": 2024,
     "sheets": ["Aplicantes Fase 1"], "estado_default": "POSTULADO", "estado_col": None},

    # 2023
    {"rel": r"5-2023\BD CONVOCATORIA 2023.xlsx", "anio": 2023,
     "sheets": ["Respuestas de formulario 1"], "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"5-2023\Fase #1 Por ciudad Jóvenes creaTIvos 2023 .xlsx", "anio": 2023,
     "sheets": ["Fase 1"], "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"5-2023\Base global información Jc 2023.xlsx", "anio": 2023,
     "sheets": ["Datos jc 2023"], "estado_default": None, "estado_col": None},

    # 2022 — AUTO (varios .xlsx, cada uno con varias hojas del mismo esquema)
    {"rel": r"4-2022\BD APLICANTES BOG 2022.xlsx", "anio": 2022, "sheets": None,
     "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"4-2022\BD APLICANTES CALI JC2022.xlsx", "anio": 2022, "sheets": None,
     "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"4-2022\BD Aplicantes MED 2022.xlsx", "anio": 2022, "sheets": None,
     "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"4-2022\BD Aplicantes ENVIGADO 2022.xlsx", "anio": 2022, "sheets": None,
     "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"4-2022\BD Inscritos BARRANQUILLA.xlsx", "anio": 2022, "sheets": None,
     "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"4-2022\Copia de BD consolidada JC2022.xlsx", "anio": 2022, "sheets": None,
     "estado_default": None, "estado_col": None},

    # 2021 — AUTO
    {"rel": r"3-2021\Copia de Aplicantes Medellín 2021.xlsx", "anio": 2021, "sheets": None,
     "estado_default": "POSTULADO", "estado_col": None},
    {"rel": r"3-2021\BD Seleccionados 2021.xlsx", "anio": 2021, "sheets": None,
     "estado_default": "SELECCIONADO", "estado_col": None},
]


def procesar():
    matcher = Matcher.desde_archivo()

    # dedup: (person_key, campo, valor, fuente_archivo) -> registro
    registros = {}
    stats_fuente = []  # por (archivo, hoja): filas, matched por metodo
    filas_totales = 0
    metodo_global = defaultdict(int)  # metodo -> nº filas
    filas_con_match = 0
    filas_sin_match = 0
    # cobertura: campo -> set de person_key (con canon) que obtuvieron el campo
    cobertura = defaultdict(set)

    for src in SOURCES:
        path = os.path.join(RAIZ_FUENTES, src["rel"])
        if not os.path.exists(path):
            print(f"[A] AVISO: no existe {path}")
            continue
        try:
            hojas = src["sheets"] if src["sheets"] else _sheetnames(path)
        except Exception as e:
            print(f"[A] AVISO: no se pudo abrir {src['rel']}: {e}")
            continue

        fuente_archivo = os.path.basename(path)
        for hoja in hojas:
            try:
                cols, filas = cargar_hoja(path, sheet=hoja)
            except Exception as e:
                print(f"[A] AVISO: no se pudo leer hoja '{hoja}' de {fuente_archivo}: {e}")
                continue
            if not cols or len(filas) < 10 or len(cols) < 4:
                continue

            campos_map, ids = construir_colmap(cols)
            tiene_id = bool(ids["cedula"] or ids["email"] or ids["nombre"])
            # AUTO: solo procesar hojas con identidad + al menos un campo socioeconómico
            if src["sheets"] is None:
                if not tiene_id or not campos_map:
                    continue
            if not tiene_id:
                continue

            estado_col = None
            if src["estado_col"]:
                estado_col = _primera_col(cols, src["estado_col"])

            n_filas = 0
            met_hoja = defaultdict(int)
            for fila in filas:
                cedula_raw = fila.get(ids["cedula"]) if ids["cedula"] else None
                email_raw = fila.get(ids["email"]) if ids["email"] else None
                nom = fila.get(ids["nombre"]) if ids["nombre"] else None
                ape = fila.get(ids["apellido"]) if ids["apellido"] else None
                nombre_full = " ".join(str(x) for x in (nom, ape) if x not in (None, "")).strip() or None

                # requiere al menos un identificador con contenido
                if not (norm_cedula(cedula_raw) or norm_email(email_raw) or norm_nombre(nombre_full)):
                    continue
                n_filas += 1

                canon, metodo = matcher.match(cedula=cedula_raw, email=email_raw, nombre=nombre_full)
                met_hoja[metodo or "sin_match"] += 1
                metodo_global[metodo or "sin_match"] += 1
                if canon:
                    filas_con_match += 1
                else:
                    filas_sin_match += 1

                cedula_fuente = norm_cedula(cedula_raw) or None
                person_key = canon or (("cf:" + cedula_fuente) if cedula_fuente
                                       else ("em:" + norm_email(email_raw)) if norm_email(email_raw)
                                       else ("nm:" + norm_nombre(nombre_full)))

                # año (por fila si la fuente no lo fija)
                anio = src["anio"]
                if anio is None and ids["anio"]:
                    raw_anio = fila.get(ids["anio"])
                    av = str(raw_anio).strip() if raw_anio is not None else ""
                    if re.fullmatch(r"\d{4}\.0", av):
                        av = av.split(".")[0]
                    if re.fullmatch(r"\d{4}", av) and 2015 <= int(av) <= 2026:
                        anio = int(av)

                pares = []  # (campo, valor)
                for campo, col in campos_map.items():
                    val = _norm_val(campo, fila.get(col))
                    if val is not None:
                        pares.append((campo, val))
                if estado_col:
                    ev = _norm_val("estado_convocatoria", fila.get(estado_col))
                    if ev:
                        pares.append(("estado_convocatoria", ev))
                if src["estado_default"]:
                    pares.append(("estado_convocatoria", src["estado_default"]))

                for campo, val in pares:
                    key = (person_key, campo, val, fuente_archivo)
                    if key in registros:
                        continue
                    registros[key] = {
                        "canon": canon,
                        "cedula_fuente": cedula_fuente,
                        "campo": campo,
                        "valor": val,
                        "fuente_archivo": fuente_archivo,
                        "hoja": hoja,
                        "anio": anio,
                        "metodo_match": metodo,
                    }
                    if canon:
                        cobertura[campo].add(canon)

            filas_totales += n_filas
            if n_filas:
                stats_fuente.append({
                    "archivo": fuente_archivo, "hoja": hoja, "filas": n_filas,
                    "campos": sorted(campos_map.keys()) + (["estado_convocatoria"] if (estado_col or src["estado_default"]) else []),
                    "metodos": dict(met_hoja),
                })

    salida = list(registros.values())
    os.makedirs(SALIDA_DIR, exist_ok=True)
    with open(SALIDA_JSON, "w", encoding="utf-8") as fh:
        json.dump(salida, fh, ensure_ascii=False, indent=1)

    escribir_resumen(salida, stats_fuente, filas_totales, filas_con_match, filas_sin_match,
                     metodo_global, cobertura, matcher)
    print(f"[A] {len(salida)} registros → {SALIDA_JSON}")
    print(f"[A] filas leídas={filas_totales}  con match={filas_con_match} "
          f"({(100.0*filas_con_match/filas_totales if filas_totales else 0):.1f}%)")


def escribir_resumen(salida, stats_fuente, filas_totales, con_match, sin_match,
                     metodo_global, cobertura, matcher):
    n_personas_roster = len(set(matcher.by_cedula.values()) | set(matcher.by_email.values()))
    personas_enriquecidas = len({r["canon"] for r in salida if r["canon"]})

    L = []
    L.append("# CLÚSTER A — Socioeconómico ampliado + convocatoria\n")
    L.append(f"Fuente raíz: `{RAIZ_FUENTES}`\n")
    L.append(f"- Registros emitidos (persona×campo×valor×fuente): **{len(salida)}**")
    L.append(f"- Filas leídas (con algún identificador): **{filas_totales}**")
    pct = (100.0 * con_match / filas_totales) if filas_totales else 0
    L.append(f"- Filas con match a persona conocida: **{con_match} ({pct:.1f}%)** · sin match: {sin_match}")
    L.append(f"- Personas del roster objetivo: ~{n_personas_roster} · personas enriquecidas por A: **{personas_enriquecidas}**\n")

    L.append("## % de match por método (sobre filas leídas)")
    for met in ("cedula", "email", "nombre", "sin_match"):
        n = metodo_global.get(met, 0)
        p = (100.0 * n / filas_totales) if filas_totales else 0
        L.append(f"- {met}: {n} ({p:.1f}%)")
    L.append("")

    L.append("## Cobertura por campo (personas distintas con canon que obtienen el campo)")
    orden = ["institucion_educativa", "grado_escolar", "promedio_colegio", "nivel_ingles",
             "ingreso_hogar", "personas_nucleo", "acceso_internet", "acceso_computador",
             "tipo_vivienda", "direccion", "barrio", "estado_convocatoria"]
    for campo in orden:
        L.append(f"- **{campo}**: {len(cobertura.get(campo, set()))}")
    otros = [c for c in cobertura if c not in orden]
    for campo in sorted(otros):
        L.append(f"- {campo}: {len(cobertura[campo])}")
    L.append("")

    L.append("## Fuentes procesadas (filas × hoja)")
    for s in stats_fuente:
        L.append(f"- `{s['archivo']}` / `{s['hoja']}` — {s['filas']} filas · campos: {', '.join(s['campos'])}")
    L.append("")

    L.append("## Ejemplos (registros con match)")
    ejemplos = [r for r in salida if r["canon"]][:12]
    for r in ejemplos:
        L.append(f"- canon={r['canon']} · {r['campo']}=\"{r['valor']}\" · {r['fuente_archivo']} "
                 f"({r['anio']}, {r['metodo_match']})")

    with open(SALIDA_RESUMEN, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L) + "\n")


if __name__ == "__main__":
    procesar()
